from openpyxl.workbook import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border
from openpyxl.styles.borders import Side
import sql_retrieve
from excel_column_number import dict_col_converter as dcc
import calendar
from dateutil.relativedelta import relativedelta
from datetime import date
import excel_formulas



def retrieve1(dict_db):
    sql = "select * from [myop\jason.walker].office_products_prod_class"
    result = sql_retrieve.getdata1(a_driver=dict_db.get('driver'), 
                                   a_server=dict_db.get('server_2'), 
                                   a_db=dict_db.get('db_playground'), 
                                   a_sql=sql)
    
    return result

def excelupdate(wb, dict_db, date_start, date_end, 
                date_ytd_start, date_trail_12_start):
    result = retrieve1(dict_db=dict_db)
    
    wb_cur = wb
    prod_class = sorted(set([r[1] for r in result]))
    ws_cur = wb_cur.create_sheet('Schedule B')
    c1 = ws_cur.cell(row=1, column=1, value='HiTouch Business Services, LLC')
    str_value = ("Schedule B - Office Products Business Unit Product Mix")
    c1 = ws_cur.cell(row=2, column=1, value=str_value)
    str_value = ("For the Period ending {d1:%B} {d1.day}, {d1.year}").format(
        d1=date_end)
    c1 = ws_cur.cell(row=3, column=1, value=str_value)    
    
    #4 Prior Months in descending order    
    month_prior_4 = []    
    month_prior_4.append(date_start)
    
    for m in range(1, 4):
        month_prior_4.append(date_start + relativedelta(months=-m))
    
    #Reverse in ascending order
    month_prior_4.reverse()
    
    row_start = 8
    
    #Product Class
    r_next = row_start
    
    for x in range(0, len(prod_class)):
        c1 = ws_cur.cell(row=r_next, column=1, value=prod_class[x])
        r_next += 1
        
    #Headers
    r_next = row_start - 1
    c1 = ws_cur.cell(row=r_next, column=1, value = 'Product Class')
    
    for c in range(2, 13, 2):
        c1 = ws_cur.cell(row=r_next, column=c, value = 'Revenues')
    
    for c in range(3, 14, 2):
        c1 = ws_cur.cell(row=r_next, column=c, value = '% of Rev')
        
    r_next = row_start - 2    
    c_next = 2
    
    for x in range(0, len(month_prior_4)):
        mth_name = calendar.month_name[month_prior_4[x].month]        
        c1 = ws_cur.cell(row=r_next, column=c_next, value = mth_name)
        c_next += 2
        
    c1 = ws_cur.cell(row=r_next, column=10, 
                     value= str(date_end.year) + ' ' + 'Year-to-Date' )
    
    c1 = ws_cur.cell(row=r_next, column=12, value='Trailing 12 Months')
    
    #Revenue Trailing 4 Months
    r_next = row_start
    
    for p in prod_class:                
        c_next = 2
        
        for c in month_prior_4:
            rev = sum([x[2] for x in result 
                       if x[0] == c 
                       if x[1] == p]) 
            c1 = ws_cur.cell(row=r_next, 
                             column=c_next, 
                             value=rev)            
            c_next += 2            
        r_next += 1
        
    #Revenue YTD                       
    r_next = row_start
    c_next = 10
    
    for p in prod_class:
        rev = sum([x[2] for x in result 
                   if date_ytd_start <= x[0] <= date_end  
                   if x[1] == p])
        c1 = ws_cur.cell(row=r_next, 
                         column=c_next, 
                         value=rev)
        r_next += 1
                
    #Revenue Trailing 12 Months                       
    r_next = row_start
    c_next = 12
    
    for p in prod_class:
        rev = sum([x[2] for x in result 
                   if date_trail_12_start <= x[0] <= date_end  
                   if x[1] == p])
        c1 = ws_cur.cell(row=r_next, 
                         column=c_next, 
                         value=rev)
        r_next += 1    
        
    #Totals
    r_last = ws_cur.max_row
    r_total = r_last + 2
    c_next = 2
    
    for c in range(c_next, 14):        
        formula1 = excel_formulas.sum_col_1(row1=row_start, 
                                            row2=r_last, 
                                            col1=c)        
        c1 = ws_cur.cell(row=r_total, column=c, value=formula1)
        
    #% of Rev
    c_next = 3
    
    for c in range(c_next, 14, 2):
        for r in range(row_start, r_last + 1):
            formula1 = excel_formulas.pct_of_total1(c, r_total, r)
            c1 = ws_cur.cell(row=r, column=c, value=formula1)
    
    #Format
    c_last = ws_cur.max_column
    
    ws_cur.merge_cells(start_row=1, end_row=1, 
                       start_column=1, end_column=c_last)
    ws_cur.merge_cells(start_row=2, end_row=2, 
                       start_column=1, end_column=c_last)
    ws_cur.merge_cells(start_row=3, end_row=3, 
                       start_column=1, end_column=c_last)    
    
    for r in range(1, 4):
        c1=ws_cur.cell(row=r, column=1)
        c1.alignment = Alignment(horizontal='center')
        c1.font = Font(bold='true')
    
    border_left = Border(left=Side(style='thin'),                         
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    
    border_right = Border(right=Side(style='thin'),                         
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))    
    
    r_next = row_start - 2
    
    for c in range(2, 13, 2):
        ws_cur.merge_cells(start_row=r_next, end_row=r_next, 
                           start_column=c, end_column=c + 1)
        c1=ws_cur.cell(row=r_next, column=c)        
        c1.alignment = Alignment(horizontal='center')
        c1.font = Font(bold='true')    
        c1.border = border_left
        c1=ws_cur.cell(row=r_next, column=c + 1)
        c1.border = border_right
    
    r_next = row_start - 1
    
    for c in range(1, c_last + 1):
        c1=ws_cur.cell(row=r_next, column=c)
        c1.font = Font(bold='true')
        c1.alignment = Alignment(horizontal='center')
        
    for c in range(2, 13, 2):
        for r in range(row_start, r_total + 1):            
            c1=ws_cur.cell(row=r, column=c)
            c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
            
    for c in range(3, 14, 2):
        for r in range(row_start, r_total + 1):            
            c1=ws_cur.cell(row=r, column=c)
            c1.number_format = '0.0%'    
    
    #Column widths
    ws_cur.column_dimensions['A'].width = 27
    
    for c in range(2, c_last + 1):
        ws_cur.column_dimensions[dcc.get(c)].width = 12.5
    
    
    #Page Setup
    ws_cur.page_setup.orientation = ws_cur.ORIENTATION_LANDSCAPE
    ws_cur.page_setup.paper_size = ws_cur.PAPERSIZE_TABLOID
    ws_cur.page_setup.fitToPage = True
    ws_cur.page_setup.fitToHeight = False
    ws_cur.page_setup.fitToWidth = 1
    ws_cur.print_options.horizontalCentered = True
    ws_cur.add_print_title(7)
    ws_cur.page_margins = PageMargins(left=.5, right=.5, top=.5, bottom=.5)
    
    #Freeze Panes
    c1 = ws_cur.cell(row=8, column=1)
    ws_cur.freeze_panes = c1            


    
    return wb_cur




if __name__ == "__main__":
    retrieve1()
    excelupdate()
    
    