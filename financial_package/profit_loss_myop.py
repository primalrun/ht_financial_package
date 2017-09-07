import sql_retrieve
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border
from openpyxl.styles.borders import Side
from copy import copy
from excel_column_number import dict_col_converter as dcc
from openpyxl.worksheet.page import PageMargins
from datetime import datetime

def retrieve_level_1(date_start, date_ytd_start, dict_db):
    sql = """
    declare @RptPeriodCur date = '{d1}'
    declare @YTDStart date = '{d2}'
    
    select
        gr.level_1,
        sum(case when gb.period = @RptPeriodCur then gb.amount else 0 end) as Amount_Mth,
        sum(case when gb.period between @YTDStart and @RptPeriodCur then gb.amount else 0 end) as Amount_YTD
    from Playground.[myop\jason.walker].gl_balance gb
        inner join Playground.[myop\jason.walker].gl_account_reporting gr
            on gb.gl_account = gr.gl_account
            and gb.company = gr.company
    where
        gb.period between @YTDStart and @RptPeriodCur
        and gr.level_1 is not null
        and gr.company = 'MYOP'
    group by
        gr.level_1
    """.format(d1=date_start, d2=date_ytd_start)
    
    gl_data = sql_retrieve.getdata1(a_driver=dict_db.get('driver'), 
                                    a_server=dict_db.get('server_2'), 
                                    a_db=dict_db.get('db_playground'), 
                                    a_sql=sql)

    return gl_data

def retrieve_level_2(date_start, date_ytd_start, dict_db):
    sql = """
    declare @RptPeriodCur date = '{d1}'
    declare @YTDStart date = '{d2}'
    
    select
        gr.level_2,
        sum(case when gb.period = @RptPeriodCur then gb.amount else 0 end) as Amount_Mth,
        sum(case when gb.period between @YTDStart and @RptPeriodCur then gb.amount else 0 end) as Amount_YTD
    from Playground.[myop\jason.walker].gl_balance gb
        inner join Playground.[myop\jason.walker].gl_account_reporting gr
            on gb.gl_account = gr.gl_account
            and gb.company = gr.company
    where
        gb.period between @YTDStart and @RptPeriodCur
        and gr.level_2 is not null
        and gr.company = 'MYOP'        
    group by
        gr.level_2
    """.format(d1=date_start, d2=date_ytd_start)
    
    gl_data = sql_retrieve.getdata1(a_driver=dict_db.get('driver'), 
                                    a_server=dict_db.get('server_2'), 
                                    a_db=dict_db.get('db_playground'), 
                                    a_sql=sql)

    return gl_data


def level_1_income(ws_cur, r_next, row_sales, category_name, level_1_month,
                         level_1_ytd, category_lookup, ):
    
    c1 = ws_cur.cell(row=r_next, column=6, value=category_name)
    c1 = ws_cur.cell(row=r_next, column=1, 
                     value=-level_1_month.get(category_lookup, 0))
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    formula1 = "=if({col1}{r_sales}=0,0,{col1}{r_cur}/{col1}{r_sales})".format(
        col1=dcc.get(1), r_sales=row_sales, r_cur=r_next)
    c1 = ws_cur.cell(row=r_next, column=2, value=formula1)
    c1.number_format = '0.0%'
    c1 = ws_cur.cell(row=r_next, column=3)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    formula1 = "=if({col1}{r_sales}=0,0,{col1}{r_cur}/{col1}{r_sales})".format(
        col1=dcc.get(3), r_sales=row_sales, r_cur=r_next)
    c1 = ws_cur.cell(row=r_next, column=4, value=formula1)
    c1.number_format = '0.0%'    
    formula1 = "={col_a}{r_cur}-{col_b}{r_cur}".format(
    col_a=dcc.get(1), col_b=dcc.get(3), r_cur=r_next)
    c1 = ws_cur.cell(row=r_next, column=5, value=formula1)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'    
    
    c1 = ws_cur.cell(row=r_next, column=7, 
                     value=-level_1_ytd.get(category_lookup, 0))
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    formula1 = "=if({col1}{r_sales}=0,0,{col1}{r_cur}/{col1}{r_sales})".format(
        col1=dcc.get(7), r_sales=row_sales, r_cur=r_next)
    c1 = ws_cur.cell(row=r_next, column=8, value=formula1)
    c1.number_format = '0.0%'
    c1 = ws_cur.cell(row=r_next, column=9)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    formula1 = "=if({col1}{r_sales}=0,0,{col1}{r_cur}/{col1}{r_sales})".format(
        col1=dcc.get(9), r_sales=row_sales, r_cur=r_next)
    c1 = ws_cur.cell(row=r_next, column=10, value=formula1)
    c1.number_format = '0.0%'    
    formula1 = "={col_a}{r_cur}-{col_b}{r_cur}".format(
    col_a=dcc.get(7), col_b=dcc.get(9), r_cur=r_next)
    c1 = ws_cur.cell(row=r_next, column=11, value=formula1)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    
    return ws_cur    

def level_1_expense(ws_cur, r_next, row_sales, category_name, level_1_month,
                         level_1_ytd, category_lookup, ):
    
    c1 = ws_cur.cell(row=r_next, column=6, value=category_name)
    c1 = ws_cur.cell(row=r_next, column=1, 
                     value=level_1_month.get(category_lookup, 0))
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    formula1 = "=if({col1}{r_sales}=0,0,{col1}{r_cur}/{col1}{r_sales})".format(
        col1=dcc.get(1), r_sales=row_sales, r_cur=r_next)
    c1 = ws_cur.cell(row=r_next, column=2, value=formula1)
    c1.number_format = '0.0%'
    c1 = ws_cur.cell(row=r_next, column=3)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    formula1 = "=if({col1}{r_sales}=0,0,{col1}{r_cur}/{col1}{r_sales})".format(
        col1=dcc.get(3), r_sales=row_sales, r_cur=r_next)
    c1 = ws_cur.cell(row=r_next, column=4, value=formula1)
    c1.number_format = '0.0%'    
    formula1 = "={col_b}{r_cur}-{col_a}{r_cur}".format(
    col_a=dcc.get(1), col_b=dcc.get(3), r_cur=r_next)
    c1 = ws_cur.cell(row=r_next, column=5, value=formula1)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'    
    
    c1 = ws_cur.cell(row=r_next, column=7, 
                     value=level_1_ytd.get(category_lookup, 0))
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    formula1 = "=if({col1}{r_sales}=0,0,{col1}{r_cur}/{col1}{r_sales})".format(
        col1=dcc.get(7), r_sales=row_sales, r_cur=r_next)
    c1 = ws_cur.cell(row=r_next, column=8, value=formula1)
    c1.number_format = '0.0%'
    c1 = ws_cur.cell(row=r_next, column=9)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    formula1 = "=if({col1}{r_sales}=0,0,{col1}{r_cur}/{col1}{r_sales})".format(
        col1=dcc.get(9), r_sales=row_sales, r_cur=r_next)
    c1 = ws_cur.cell(row=r_next, column=10, value=formula1)
    c1.number_format = '0.0%'    
    formula1 = "={col_b}{r_cur}-{col_a}{r_cur}".format(
    col_a=dcc.get(7), col_b=dcc.get(9), r_cur=r_next)
    c1 = ws_cur.cell(row=r_next, column=11, value=formula1)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'    

    return ws_cur


def level_2_income(ws_cur, r_next, row_sales, category_name, level_2_month,
                         level_2_ytd, category_lookup, ):
    
    c1 = ws_cur.cell(row=r_next, column=6, value=category_name)
    c1 = ws_cur.cell(row=r_next, column=1, 
                     value=-level_2_month.get(category_lookup, 0))
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    formula1 = "=if({col1}{r_sales}=0,0,{col1}{r_cur}/{col1}{r_sales})".format(
        col1=dcc.get(1), r_sales=row_sales, r_cur=r_next)
    c1 = ws_cur.cell(row=r_next, column=2, value=formula1)
    c1.number_format = '0.0%'
    c1 = ws_cur.cell(row=r_next, column=3)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    formula1 = "=if({col1}{r_sales}=0,0,{col1}{r_cur}/{col1}{r_sales})".format(
        col1=dcc.get(3), r_sales=row_sales, r_cur=r_next)
    c1 = ws_cur.cell(row=r_next, column=4, value=formula1)
    c1.number_format = '0.0%'    
    formula1 = "={col_a}{r_cur}-{col_b}{r_cur}".format(
    col_a=dcc.get(1), col_b=dcc.get(3), r_cur=r_next)
    c1 = ws_cur.cell(row=r_next, column=5, value=formula1)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'    
    
    c1 = ws_cur.cell(row=r_next, column=7, 
                     value=-level_2_ytd.get(category_lookup, 0))
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    formula1 = "=if({col1}{r_sales}=0,0,{col1}{r_cur}/{col1}{r_sales})".format(
        col1=dcc.get(7), r_sales=row_sales, r_cur=r_next)
    c1 = ws_cur.cell(row=r_next, column=8, value=formula1)
    c1.number_format = '0.0%'
    c1 = ws_cur.cell(row=r_next, column=9)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    formula1 = "=if({col1}{r_sales}=0,0,{col1}{r_cur}/{col1}{r_sales})".format(
        col1=dcc.get(9), r_sales=row_sales, r_cur=r_next)
    c1 = ws_cur.cell(row=r_next, column=10, value=formula1)
    c1.number_format = '0.0%'    
    formula1 = "={col_a}{r_cur}-{col_b}{r_cur}".format(
    col_a=dcc.get(7), col_b=dcc.get(9), r_cur=r_next)
    c1 = ws_cur.cell(row=r_next, column=11, value=formula1)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    
    return ws_cur    


def level_2_expense(ws_cur, r_next, row_sales, category_name, level_2_month,
                         level_2_ytd, category_lookup, ):
    
    c1 = ws_cur.cell(row=r_next, column=6, value=category_name)
    c1 = ws_cur.cell(row=r_next, column=1, 
                     value=level_2_month.get(category_lookup, 0))
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    formula1 = "=if({col1}{r_sales}=0,0,{col1}{r_cur}/{col1}{r_sales})".format(
        col1=dcc.get(1), r_sales=row_sales, r_cur=r_next)
    c1 = ws_cur.cell(row=r_next, column=2, value=formula1)
    c1.number_format = '0.0%'
    c1 = ws_cur.cell(row=r_next, column=3)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    formula1 = "=if({col1}{r_sales}=0,0,{col1}{r_cur}/{col1}{r_sales})".format(
        col1=dcc.get(3), r_sales=row_sales, r_cur=r_next)
    c1 = ws_cur.cell(row=r_next, column=4, value=formula1)
    c1.number_format = '0.0%'    
    formula1 = "={col_b}{r_cur}-{col_a}{r_cur}".format(
    col_a=dcc.get(1), col_b=dcc.get(3), r_cur=r_next)
    c1 = ws_cur.cell(row=r_next, column=5, value=formula1)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'    
    
    c1 = ws_cur.cell(row=r_next, column=7, 
                     value=level_2_ytd.get(category_lookup, 0))
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    formula1 = "=if({col1}{r_sales}=0,0,{col1}{r_cur}/{col1}{r_sales})".format(
        col1=dcc.get(7), r_sales=row_sales, r_cur=r_next)
    c1 = ws_cur.cell(row=r_next, column=8, value=formula1)
    c1.number_format = '0.0%'
    c1 = ws_cur.cell(row=r_next, column=9)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    formula1 = "=if({col1}{r_sales}=0,0,{col1}{r_cur}/{col1}{r_sales})".format(
        col1=dcc.get(9), r_sales=row_sales, r_cur=r_next)
    c1 = ws_cur.cell(row=r_next, column=10, value=formula1)
    c1.number_format = '0.0%'    
    formula1 = "={col_b}{r_cur}-{col_a}{r_cur}".format(
    col_a=dcc.get(7), col_b=dcc.get(9), r_cur=r_next)
    c1 = ws_cur.cell(row=r_next, column=11, value=formula1)
    c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'    

    return ws_cur


def build_consolidated_p_and_l(wb, date_start, date_ytd_start, 
                               dict_db, date_end):
    wb_cur = wb    
    level_1 = retrieve_level_1(date_start, date_ytd_start, dict_db)
    level_2 = retrieve_level_2(date_start, date_ytd_start, dict_db)
    
    level_1_month = {}
    level_1_ytd = {}
    
    for x in range(0, len(level_1)):
        level_1_month[level_1[x][0]] = level_1[x][1] 
        level_1_ytd[level_1[x][0]] = level_1[x][2]
    
    level_2_month = {}
    level_2_ytd = {}
    
    for x in range(0, len(level_2)):
        level_2_month[level_2[x][0]] = level_2[x][1] 
        level_2_ytd[level_2[x][0]] = level_2[x][2]    
    
    ws_cur = wb_cur.create_sheet('P&L_MYOP')
        
    c1 = ws_cur.cell(row=1, column=1, value='P&L - MYOP')
    c1 = ws_cur.cell(row=2, column=1, value='Income Statement')
    
    report_date = '{d1:%B} {d1.day}, {d1.year}'.format(
        d1=date_end)
    c1 = ws_cur.cell(row=3, column=1, value=report_date)    
    
    row_header_1 = 5
    border_tb = Border(top=Side(style='thin'),                         
                     bottom=Side(style='thin'))
    border_tb2 = Border(top=Side(style='thin'),                         
                     bottom=Side(style='double', ))    
    border_all = Border(top=Side(style='thin'),                         
                     bottom=Side(style='thin'),
                     left=Side(style='thin'),
                     right=Side(style='thin'))    
    
    for r in range(1, 4):
        ws_cur.merge_cells(start_row=r,
                           end_row=r,
                           start_column=1,
                           end_column=11)
        c1 = ws_cur.cell(row=r, column=1)
        c1.font = Font(bold='true')
        c1.alignment = Alignment(horizontal='center')
        
    c1 = ws_cur.cell(row=row_header_1, column=1, value='Month')
    ws_cur.merge_cells(start_row=row_header_1,
                       end_row=row_header_1,
                       start_column=1,
                       end_column=5)
    c1.font = Font(bold='true')
    c1.alignment = Alignment(horizontal='center')
    
    c1 = ws_cur.cell(row=row_header_1, column=7, value='Year to Date')
    ws_cur.merge_cells(start_row=row_header_1,
                       end_row=row_header_1,
                       start_column=7,
                       end_column=11)
    c1.font = Font(bold='true')
    c1.alignment = Alignment(horizontal='center')
        
    row_header_2 = 6
    c1 = ws_cur.cell(row=row_header_2, column=1, value='Actual')
    c1 = ws_cur.cell(row=row_header_2, column=2, value='% Sales')
    c1 = ws_cur.cell(row=row_header_2, column=3, value='Budget')
    c1 = ws_cur.cell(row=row_header_2, column=4, value='% Sales')
    c1 = ws_cur.cell(row=row_header_2, column=5, value='Variance')
    
    for c in range(1, 6):
        c1 = ws_cur.cell(row=row_header_2, column=c)
        c1.fill = PatternFill(start_color='1e90ff', end_color='1e90ff', 
                              fill_type='solid')
        c1.font = Font(bold='true', color='f8f8ff')
        c1.alignment = Alignment(horizontal='center')
    
    c1 = ws_cur.cell(row=row_header_2, column=7, value='Actual')
    c1 = ws_cur.cell(row=row_header_2, column=8, value='% Sales')
    c1 = ws_cur.cell(row=row_header_2, column=9, value='Budget')
    c1 = ws_cur.cell(row=row_header_2, column=10, value='% Sales')
    c1 = ws_cur.cell(row=row_header_2, column=11, value='Variance')
    
    for c in range(7, 12):
        c1 = ws_cur.cell(row=row_header_2, column=c)
        c1.fill = PatternFill(start_color='1e90ff', end_color='1e90ff', 
                              fill_type='solid')
        c1.font = Font(bold='true', color='f8f8ff')
        c1.alignment = Alignment(horizontal='center')
    
    for c in [x for x in range(1, 12) if x != 6]:
        c1 = ws_cur.cell(row=row_header_2, column=c)
        c1.border = border_all
    
    #------Revenue------
    r_next = row_header_2 + 1    
    row_sales = r_next    
    ws_cur = level_1_income(ws_cur, r_next, row_sales, 'Total Revenue',
                         level_1_month, level_1_ytd, 'Sales')
    
    for c in range(1, 12):
        c1=ws_cur.cell(row=r_next, column=c)
        c1.font = Font(bold='true')
    
    #------Cost------
    r_next += 1
    ws_cur = level_1_expense(ws_cur, r_next, row_sales, 'Cost of Sales',
                         level_1_month, level_1_ytd, 'Cost')    
    
    for c in range(1, 12):
        c1=ws_cur.cell(row=r_next, column=c)
        c1.font = Font(bold='true')
    
    #------Gross Margin------
    r_next += 1
    row_gm = r_next
    c1 = ws_cur.cell(row=r_next, column=6, value = 'Gross Margin')    
    
    #calculation
    for c in [x for x in range(1, 12) if x in [1, 3, 7, 9]]:
        formula1 = "={col_letter}{r1}-{col_letter}{r2}".format(
            col_letter = dcc.get(c), r1=r_next - 2, r2=r_next - 1)        
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    
    #% sales    
    for c in [x for x in range(2, 11, 2) if x != 6]:
        formula1 = (""
            "=if({col1}{r_sales}=0,0,{col1}{r_cur}/{col1}{r_sales})").format(
                col1=dcc.get(c - 1), r_sales=row_sales, r_cur=r_next)
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)    
        c1.number_format = '0.0%'
    
    #variance    
    for c in range(5, 12, 6):
        formula1 = "={col_a}{r_cur}-{col_b}{r_cur}".format(
            col_a=dcc.get(c - 4), col_b=dcc.get(c - 2), r_cur=r_next)
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)    
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'                         
    
    #border            
    for c in [x for x in range(1, 12) if x != 6]:
        c1=ws_cur.cell(row=r_next, column=c)
        c1.border = border_tb
    
    for c in range(1, 12):
        c1=ws_cur.cell(row=r_next, column=c)
        c1.font = Font(bold='true')
            
    #------Adjustments to Margin------
    r_next += 2
    c1 = ws_cur.cell(row=r_next, column=6, value = 'Adjustments to Margin:')
    c1.font = Font(bold='true')
    
        #------Rebates------
    r_next += 1
    ws_cur = level_2_income(ws_cur, r_next, row_sales, 'Rebates',
                         level_2_month, level_2_ytd, 'Rebates')
    
        #------Cash Discounts------
    r_next += 1
    ws_cur = level_2_income(ws_cur, r_next, row_sales, 'Cash Discounts',
                         level_2_month, level_2_ytd, 'Cash Discounts')    
    
        #------Customer Discounts------
    r_next += 1
    ws_cur = level_2_income(ws_cur, r_next, row_sales, 'Customer Discounts',
                         level_2_month, level_2_ytd, 'Customer Discounts')
    
        #------Cost of Goods Adjustments------
    r_next += 1
    ws_cur = level_2_income(ws_cur, r_next, row_sales, 'Cost of Goods Adjustments',
                         level_2_month, level_2_ytd, 'Cost of Goods Adjustments')    
    
        #------Other------
    r_next += 1
    ws_cur = level_2_income(ws_cur, r_next, row_sales, 'Other',
                         level_2_month, level_2_ytd, 'Other Adjustments to Margin')    
        
    #------Total Adjustments to Margin------
    r_next += 1
    c1 = ws_cur.cell(row=r_next, column=6, value = 'Total Adjustments to Margin')    
    
    #calculation
    for c in [x for x in range(1, 12) if x in [1, 3, 7, 9]]:
        formula1 = "=sum({col_letter}{r1}:{col_letter}{r2})".format(
            col_letter = dcc.get(c), r1=r_next - 5, r2=r_next - 1)        
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    
    #% sales    
    for c in [x for x in range(2, 11, 2) if x != 6]:
        formula1 = (""
            "=if({col1}{r_sales}=0,0,{col1}{r_cur}/{col1}{r_sales})").format(
                col1=dcc.get(c - 1), r_sales=row_sales, r_cur=r_next)
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)    
        c1.number_format = '0.0%'
    
    #variance    
    for c in range(5, 12, 6):
        formula1 = "={col_a}{r_cur}-{col_b}{r_cur}".format(
            col_a=dcc.get(c - 4), col_b=dcc.get(c - 2), r_cur=r_next)
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)    
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'                         
    
    #border            
    for c in [x for x in range(1, 12) if x != 6]:
        c1=ws_cur.cell(row=r_next, column=c)
        c1.border = border_tb    
    
    for c in range(1, 12):
        c1=ws_cur.cell(row=r_next, column=c)
        c1.font = Font(bold='true')
            
    #------Adjusted Gross Margin------
    r_next += 1
    c1 = ws_cur.cell(row=r_next, column=6, value = 'Adjusted Gross Margin')    
    
    #calculation
    for c in [x for x in range(1, 12) if x in [1, 3, 7, 9]]:
        formula1 = "={col_letter}{r1}+{col_letter}{r2}".format(
            col_letter = dcc.get(c), r1=r_next - 9, r2=r_next - 1)        
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    
    #% sales    
    for c in [x for x in range(2, 11, 2) if x != 6]:
        formula1 = (""
            "=if({col1}{r_sales}=0,0,{col1}{r_cur}/{col1}{r_sales})").format(
                col1=dcc.get(c - 1), r_sales=row_sales, r_cur=r_next)
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)    
        c1.number_format = '0.0%'
    
    #variance    
    for c in range(5, 12, 6):
        formula1 = "={col_a}{r_cur}-{col_b}{r_cur}".format(
            col_a=dcc.get(c - 4), col_b=dcc.get(c - 2), r_cur=r_next)
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)    
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'                         
    
    #border            
    for c in [x for x in range(1, 12) if x != 6]:
        c1=ws_cur.cell(row=r_next, column=c)
        c1.border = border_tb    
    
    for c in range(1, 12):
        c1=ws_cur.cell(row=r_next, column=c)
        c1.font = Font(bold='true')
        
        #------Commissions------
    r_next += 2
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Commissions',
                         level_2_month, level_2_ytd, 'Commissions')            
        
    #------Total Commissions------
    r_next += 1
    c1 = ws_cur.cell(row=r_next, column=6, value = 'Total Commissions')    
    
    #calculation
    for c in [x for x in range(1, 12) if x in [1, 3, 7, 9]]:
        formula1 = "=sum({col_letter}{r1}:{col_letter}{r2})".format(
            col_letter = dcc.get(c), r1=r_next - 1, r2=r_next - 1)        
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    
    #% sales    
    for c in [x for x in range(2, 11, 2) if x != 6]:
        formula1 = (""
            "=if({col1}{r_sales}=0,0,{col1}{r_cur}/{col1}{r_sales})").format(
                col1=dcc.get(c - 1), r_sales=row_sales, r_cur=r_next)
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)    
        c1.number_format = '0.0%'
    
    #variance    
    for c in range(5, 12, 6):
        formula1 = "={col_b}{r_cur}-{col_a}{r_cur}".format(
            col_a=dcc.get(c - 4), col_b=dcc.get(c - 2), r_cur=r_next)
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)    
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'                         
    
    #border            
    for c in [x for x in range(1, 12) if x != 6]:
        c1=ws_cur.cell(row=r_next, column=c)
        c1.border = border_tb    
    
    for c in range(1, 12):
        c1=ws_cur.cell(row=r_next, column=c)
        c1.font = Font(bold='true')    
    
    
    #% To Margin
    r_next += 1
    c1=ws_cur.cell(row=r_next, column=6, value='% To Margin')    
    
    for c in [x for x in range(1, 12) if x in [1, 3, 7, 9]]:
        formula1 = (""
            "=if({col1}{r_gm}=0,0,{col1}{r_cur}/{col1}{r_gm})").format(
                col1=dcc.get(c), r_gm=row_gm, r_cur=r_next - 1)    
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)    
        c1.number_format = '0.0%'
        
    for c in range(1, 12):
        c1 = ws_cur.cell(row=r_next, column=c)
        c1.font = Font(italic='true')
        
    #------Other Personnel Expenses------
    r_next += 2
    c1 = ws_cur.cell(row=r_next, column=6, value = 'Other Personnel Expenses:')
    c1.font = Font(bold='true')        
    
        #------Salaries------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Salaries',
                         level_2_month, level_2_ytd, 'Salaries')    
    
        #------Bonuses------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Bonuses',
                         level_2_month, level_2_ytd, 'Bonuses')
    
        #------Benefits------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Benefits',
                         level_2_month, level_2_ytd, 'Benefits')
    
        #------Payroll Taxes------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Payroll Taxes',
                         level_2_month, level_2_ytd, 'Payroll Taxes')            
    
    #------Total Other Personnel Expenses------
    r_next += 1
    c1 = ws_cur.cell(row=r_next, column=6, value = 'Total Other Personnel Expenses')    
    
    #calculation
    for c in [x for x in range(1, 12) if x in [1, 3, 7, 9]]:
        formula1 = "=sum({col_letter}{r1}:{col_letter}{r2})".format(
            col_letter = dcc.get(c), r1=r_next - 4, r2=r_next - 1)        
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    
    #% sales    
    for c in [x for x in range(2, 11, 2) if x != 6]:
        formula1 = (""
            "=if({col1}{r_sales}=0,0,{col1}{r_cur}/{col1}{r_sales})").format(
                col1=dcc.get(c - 1), r_sales=row_sales, r_cur=r_next)
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)    
        c1.number_format = '0.0%'
    
    #variance    
    for c in range(5, 12, 6):
        formula1 = "={col_b}{r_cur}-{col_a}{r_cur}".format(
            col_a=dcc.get(c - 4), col_b=dcc.get(c - 2), r_cur=r_next)
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)    
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'                         
    
    #border            
    for c in [x for x in range(1, 12) if x != 6]:
        c1=ws_cur.cell(row=r_next, column=c)
        c1.border = border_tb    
    
    for c in range(1, 12):
        c1=ws_cur.cell(row=r_next, column=c)
        c1.font = Font(bold='true')    
    
    #------Other Operating Expenses------
    r_next += 2
    c1 = ws_cur.cell(row=r_next, column=6, value = 'Other Operating Expenses:')
    c1.font = Font(bold='true')        
    
        #------Travel------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Travel',
                         level_2_month, level_2_ytd, 'Travel')
    
        #------Meals & Entertainment------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Meals & Entertainment',
                         level_2_month, level_2_ytd, 'Meals & Entertainment')

        #------Facility Rent------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Facility Rent',
                         level_2_month, level_2_ytd, 'Facility Rent')

        #------Utilities------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Utilities',
                         level_2_month, level_2_ytd, 'Utilities')

        #------Facility Maintenance------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Facility Maintenance',
                         level_2_month, level_2_ytd, 'Facility Maintenance')

        #------Fleet Fuel------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Fleet Fuel',
                         level_2_month, level_2_ytd, 'Fleet Fuel')

        #------Fleet Repair & Maintenance------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Fleet Repair & Maintenance',
                         level_2_month, level_2_ytd, 'Fleet Repair & Maintenance')

        #------Vehicle Rent Expense------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Vehicle Rent Expense',
                         level_2_month, level_2_ytd, 'Vehicle Rent Expense')

        #------Outsourced Delivery------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Outsourced Delivery',
                         level_2_month, level_2_ytd, 'Outsourced Delivery')

        #------Outbound Freight------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Outbound Freight',
                         level_2_month, level_2_ytd, 'Outbound Freight')

        #------Outbound Freight Rebates------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Outbound Freight Rebates',
                         level_2_month, level_2_ytd, 'Outbound Freight Rebates')

        #------Equipment Rental------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Equipment Rental',
                         level_2_month, level_2_ytd, 'Equipment Rental')

        #------MRO Supplies------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'MRO Supplies',
                         level_2_month, level_2_ytd, 'MRO Supplies')

        #------Office Freight & Postage------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Office Freight & Postage',
                         level_2_month, level_2_ytd, 'Office Freight & Postage')

        #------Outside Services------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Outside Services',
                         level_2_month, level_2_ytd, 'Outside Services')

        #------IT Services & Maintenance------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'IT Services & Maintenance',
                         level_2_month, level_2_ytd, 'IT Services & Maintenance')

        #------Telecom------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Telecom Expense',
                         level_2_month, level_2_ytd, 'Telecom Expense')

        #------Business Insurance------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Business Insurance',
                         level_2_month, level_2_ytd, 'Business Insurance')

        #------Training, Dues & Subscriptions------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Training, Dues & Subscriptions',
                         level_2_month, level_2_ytd, 'Training, Dues & Subscriptions')

        #------Advertising & Marketing------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Advertising & Marketing',
                         level_2_month, level_2_ytd, 'Advertising & Marketing')

        #------Diversity Partner Fees------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Diversity Partner Fees',
                         level_2_month, level_2_ytd, 'Diversity Partner Fees')

        #------e-Commerce Fees------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'E-Commerce Fees',
                         level_2_month, level_2_ytd, 'e-Commerce Fees')

        #------Professional Fees------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Professional Fees',
                         level_2_month, level_2_ytd, 'Professional Fees')

        #------Bank Charges------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Bank Charges',
                         level_2_month, level_2_ytd, 'Bank Charges')

        #------Bad Debt Expense------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Bad Debt Expense',
                         level_2_month, level_2_ytd, 'Bad Debt Expense')

        #------Other Taxes------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Other Taxes',
                         level_2_month, level_2_ytd, 'Other Taxes')

        #------Other Expenses------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Other Expenses',
                         level_2_month, level_2_ytd, 'Other Expenses')            
    
    #------Total Other Operating Expenses------
    r_next += 1
    c1 = ws_cur.cell(row=r_next, column=6, value = 'Total Other Operating Expenses')    
    
    #calculation
    for c in [x for x in range(1, 12) if x in [1, 3, 7, 9]]:
        formula1 = "=sum({col_letter}{r1}:{col_letter}{r2})".format(
            col_letter = dcc.get(c), r1=r_next - 27, r2=r_next - 1)        
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    
    #% sales    
    for c in [x for x in range(2, 11, 2) if x != 6]:
        formula1 = (""
            "=if({col1}{r_sales}=0,0,{col1}{r_cur}/{col1}{r_sales})").format(
                col1=dcc.get(c - 1), r_sales=row_sales, r_cur=r_next)
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)    
        c1.number_format = '0.0%'
    
    #variance    
    for c in range(5, 12, 6):
        formula1 = "={col_b}{r_cur}-{col_a}{r_cur}".format(
            col_a=dcc.get(c - 4), col_b=dcc.get(c - 2), r_cur=r_next)
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)    
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'                         
    
    #border            
    for c in [x for x in range(1, 12) if x != 6]:
        c1=ws_cur.cell(row=r_next, column=c)
        c1.border = border_tb    
    
    for c in range(1, 12):
        c1=ws_cur.cell(row=r_next, column=c)
        c1.font = Font(bold='true')        
    
    #------Total Operating Expenses------
    r_next += 1
    c1 = ws_cur.cell(row=r_next, column=6, value = 'Total Operating Expenses')    
    
    #calculation
    for c in [x for x in range(1, 12) if x in [1, 3, 7, 9]]:
        formula1 = "={col_letter}{r1}+{col_letter}{r2}+{col_letter}{r3}".format(
            col_letter = dcc.get(c), r1=r_next - 39, 
            r2=r_next - 31, r3=r_next - 1)        
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    
    #% sales    
    for c in [x for x in range(2, 11, 2) if x != 6]:
        formula1 = (""
            "=if({col1}{r_sales}=0,0,{col1}{r_cur}/{col1}{r_sales})").format(
                col1=dcc.get(c - 1), r_sales=row_sales, r_cur=r_next)
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)    
        c1.number_format = '0.0%'
    
    #variance    
    for c in range(5, 12, 6):
        formula1 = "={col_b}{r_cur}-{col_a}{r_cur}".format(
            col_a=dcc.get(c - 4), col_b=dcc.get(c - 2), r_cur=r_next)
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)    
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'                         
    
    #border            
    for c in [x for x in range(1, 12) if x != 6]:
        c1=ws_cur.cell(row=r_next, column=c)
        c1.border = border_tb    
    
    for c in range(1, 12):
        c1=ws_cur.cell(row=r_next, column=c)
        c1.font = Font(bold='true')    
    
    #------Other Income / (Expense)------
    r_next += 2
    c1 = ws_cur.cell(row=r_next, column=6, value = 'Other Income / (Expense):')
    c1.font = Font(bold='true')        
    
        #------Corporate Allocation------
    r_next += 1
    ws_cur = level_2_income(ws_cur, r_next, row_sales, 'Corporate Allocation',
                         level_2_month, level_2_ytd, 'Corporate Allocation')    
    
        #------Deferred Partner Revenue------    
    r_next += 1
    ws_cur = level_2_income(ws_cur, r_next, row_sales, 'Deferred Partner Revenue',
                         level_2_month, level_2_ytd, 'Deferred Partner Revenue')    

        #------Excluded Expenses------
    r_next += 1
    ws_cur = level_2_income(ws_cur, r_next, row_sales, 'Excluded Expenses',
                         level_2_month, level_2_ytd, 'Excluded Expenses')    

        #------Misc. Income------
    r_next += 1
    ws_cur = level_2_income(ws_cur, r_next, row_sales, 'Misc. Income',
                         level_2_month, level_2_ytd, 'Misc. Income')    

        #------Gain/(Loss) on Fixed Asset Disposal------
    r_next += 1
    ws_cur = level_2_income(ws_cur, r_next, row_sales, 'Gain / (Loss) Fixed Asset Disposal',
                         level_2_month, level_2_ytd, 'Gain / (Loss) Fixed Asset Disposal')    
    
    #------Total Other Income / (Expense)------
    r_next += 1
    c1 = ws_cur.cell(row=r_next, column=6, value = 'Total Other Income / (Expense)')    
    
    #calculation
    for c in [x for x in range(1, 12) if x in [1, 3, 7, 9]]:
        formula1 = "=sum({col_letter}{r1}:{col_letter}{r2})".format(
            col_letter = dcc.get(c), r1=r_next - 5, r2=r_next - 1)        
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    
    #% sales    
    for c in [x for x in range(2, 11, 2) if x != 6]:
        formula1 = (""
            "=if({col1}{r_sales}=0,0,{col1}{r_cur}/{col1}{r_sales})").format(
                col1=dcc.get(c - 1), r_sales=row_sales, r_cur=r_next)
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)    
        c1.number_format = '0.0%'
    
    #variance    
    for c in range(5, 12, 6):
        formula1 = "={col_a}{r_cur}-{col_b}{r_cur}".format(
            col_a=dcc.get(c - 4), col_b=dcc.get(c - 2), r_cur=r_next)
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)    
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'                         
    
    #border            
    for c in [x for x in range(1, 12) if x != 6]:
        c1=ws_cur.cell(row=r_next, column=c)
        c1.border = border_tb    
    
    for c in range(1, 12):
        c1=ws_cur.cell(row=r_next, column=c)
        c1.font = Font(bold='true')
        
    #-----EBITDA------
    r_next += 2
    c1 = ws_cur.cell(row=r_next, column=6, value = 'EBITDA')    
    
    #calculation
    for c in [x for x in range(1, 12) if x in [1, 3, 7, 9]]:
        formula1 = "={col_letter}{r1}-{col_letter}{r2}+{col_letter}{r3}".format(
            col_letter = dcc.get(c), r1=r_next - 52, r2=r_next - 10, 
            r3= r_next - 2)        
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    
    #% sales    
    for c in [x for x in range(2, 11, 2) if x != 6]:
        formula1 = (""
            "=if({col1}{r_sales}=0,0,{col1}{r_cur}/{col1}{r_sales})").format(
                col1=dcc.get(c - 1), r_sales=row_sales, r_cur=r_next)
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)    
        c1.number_format = '0.0%'
    
    #variance    
    for c in range(5, 12, 6):
        formula1 = "={col_a}{r_cur}-{col_b}{r_cur}".format(
            col_a=dcc.get(c - 4), col_b=dcc.get(c - 2), r_cur=r_next)
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)    
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'                         
    
    #border            
    for c in [x for x in range(1, 12) if x != 6]:
        c1=ws_cur.cell(row=r_next, column=c)
        c1.border = border_tb    
    
    for c in range(1, 12):
        c1=ws_cur.cell(row=r_next, column=c)
        c1.font = Font(bold='true')        
        c1.fill = PatternFill(start_color='ffd700', end_color='ffd700', 
                              fill_type='solid')
        
        #------Interest Expense------
    r_next += 2
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Interest Expense',
                         level_2_month, level_2_ytd, 'Interest Expense')

        #------Depreciation------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Depreciation',
                         level_2_month, level_2_ytd, 'Depreciation')

        #------Amortization------
    r_next += 1
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'Amortization',
                         level_2_month, level_2_ytd, 'Amortization')            
    
    #------Net Income Before Taxes------
    r_next += 1
    c1 = ws_cur.cell(row=r_next, column=6, value = 'Net Income Before Taxes')    
    
    #calculation
    for c in [x for x in range(1, 12) if x in [1, 3, 7, 9]]:
        formula1 = ("={col_letter}{r3}-"
                    "sum({col_letter}{r1}:{col_letter}{r2})").format(
                        col_letter = dcc.get(c), r1=r_next - 3, r2=r_next - 1, 
                        r3=r_next - 5)        
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    
    #% sales    
    for c in [x for x in range(2, 11, 2) if x != 6]:
        formula1 = (""
            "=if({col1}{r_sales}=0,0,{col1}{r_cur}/{col1}{r_sales})").format(
                col1=dcc.get(c - 1), r_sales=row_sales, r_cur=r_next)
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)    
        c1.number_format = '0.0%'
    
    #variance    
    for c in range(5, 12, 6):
        formula1 = "={col_a}{r_cur}-{col_b}{r_cur}".format(
            col_a=dcc.get(c - 4), col_b=dcc.get(c - 2), r_cur=r_next)
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)    
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'                         
    
    #border            
    for c in [x for x in range(1, 12) if x != 6]:
        c1=ws_cur.cell(row=r_next, column=c)
        c1.border = border_tb    
    
    for c in range(1, 12):
        c1=ws_cur.cell(row=r_next, column=c)
        c1.font = Font(bold='true')                    
    
        #------State Income Taxes------
    r_next += 2
    ws_cur = level_2_expense(ws_cur, r_next, row_sales, 'State Income Taxes',
                         level_2_month, level_2_ytd, 'State Income Taxes')            
    
    #------Net Income------
    r_next += 1
    c1 = ws_cur.cell(row=r_next, column=6, value = 'Net Income')    
    
    #calculation
    for c in [x for x in range(1, 12) if x in [1, 3, 7, 9]]:
        formula1 = ("={col_letter}{r1}-{col_letter}{r2}").format(
                        col_letter = dcc.get(c), r1=r_next - 3, r2=r_next - 1)
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    
    #% sales    
    for c in [x for x in range(2, 11, 2) if x != 6]:
        formula1 = (""
            "=if({col1}{r_sales}=0,0,{col1}{r_cur}/{col1}{r_sales})").format(
                col1=dcc.get(c - 1), r_sales=row_sales, r_cur=r_next)
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)    
        c1.number_format = '0.0%'
    
    #variance    
    for c in range(5, 12, 6):
        formula1 = "={col_a}{r_cur}-{col_b}{r_cur}".format(
            col_a=dcc.get(c - 4), col_b=dcc.get(c - 2), r_cur=r_next)
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)    
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'                         
    
    #border            
    for c in [x for x in range(1, 12) if x != 6]:
        c1=ws_cur.cell(row=r_next, column=c)
        c1.border = border_tb2    
    
    for c in range(1, 12):
        c1=ws_cur.cell(row=r_next, column=c)
        c1.font = Font(bold='true')                    
    
    #------Excluded Expenses------
    r_next += 2
    c1 = ws_cur.cell(row=r_next, column=6, value = 'EXCLUDED EXPENSES')    
    
    #calculation
    for c in [x for x in range(1, 12) if x in [1, 3, 7, 9]]:
        formula1 = "=-{col_letter}{r1}".format(
                        col_letter = dcc.get(c), r1=r_next - 15)
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    
    #% sales    
    for c in [x for x in range(2, 11, 2) if x != 6]:
        formula1 = (""
            "=if({col1}{r_sales}=0,0,{col1}{r_cur}/{col1}{r_sales})").format(
                col1=dcc.get(c - 1), r_sales=row_sales, r_cur=r_next)
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)    
        c1.number_format = '0.0%'
    
    #variance    
    for c in range(5, 12, 6):
        formula1 = "={col_a}{r_cur}-{col_b}{r_cur}".format(
            col_a=dcc.get(c - 4), col_b=dcc.get(c - 2), r_cur=r_next)
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)    
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'                         
    
    #border            
    for c in [x for x in range(1, 12) if x != 6]:
        c1=ws_cur.cell(row=r_next, column=c)
        c1.border = border_tb    
    
    for c in range(1, 12):
        c1=ws_cur.cell(row=r_next, column=c)
        c1.font = Font(bold='true')        
        c1.fill = PatternFill(start_color='ffd700', end_color='ffd700', 
                              fill_type='solid')

    #------Adjusted EBITDA------
    r_next += 2
    c1 = ws_cur.cell(row=r_next, column=6, value = 'ADJUSTED EBITDA')    
    
    #calculation
    for c in [x for x in range(1, 12) if x in [1, 3, 7, 9]]:
        formula1 = "={col_letter}{r1}+{col_letter}{r2}".format(
                        col_letter = dcc.get(c), r1=r_next - 12, 
                        r2=r_next - 2)
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    
    #% sales    
    for c in [x for x in range(2, 11, 2) if x != 6]:
        formula1 = (""
            "=if({col1}{r_sales}=0,0,{col1}{r_cur}/{col1}{r_sales})").format(
                col1=dcc.get(c - 1), r_sales=row_sales, r_cur=r_next)
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)    
        c1.number_format = '0.0%'
    
    #variance    
    for c in range(5, 12, 6):
        formula1 = "={col_a}{r_cur}-{col_b}{r_cur}".format(
            col_a=dcc.get(c - 4), col_b=dcc.get(c - 2), r_cur=r_next)
        c1 = ws_cur.cell(row=r_next, column=c, value = formula1)    
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'                         
    
    #border            
    for c in [x for x in range(1, 12) if x != 6]:
        c1=ws_cur.cell(row=r_next, column=c)
        c1.border = border_tb    
    
    for c in range(1, 12):
        c1=ws_cur.cell(row=r_next, column=c)
        c1.font = Font(bold='true')        
        c1.fill = PatternFill(start_color='ffd700', end_color='ffd700', 
                              fill_type='solid')


    
    #Change font size and name
    for r in range(1, ws_cur.max_row + 1):
        for c in range(1, ws_cur.max_column + 1):        
            c1 = ws_cur.cell(row=r, column=c)
            c1.font = Font(bold=c1.font.bold,
                           italic=c1.font.italic,
                           name='Arial',
                           size=8,
                           color=c1.font.color)
            
    #Row height
    for r in range(1, ws_cur.max_row + 1):
        ws_cur.row_dimensions[r].height = 11.5        
    
    #Column width
    for x in [x for x in range(1, 12) if x in [1, 3, 5, 7, 9, 11]]:
        ws_cur.column_dimensions[dcc.get(x)].width=11
         
    for x in [x for x in range(1, 12) if x in [2, 4, 8, 10]]:
        ws_cur.column_dimensions[dcc.get(x)].width=6.5
        
    ws_cur.column_dimensions[dcc.get(6)].width=27    
                
        
    #Page Setup
    ws_cur.page_setup.orientation = ws_cur.ORIENTATION_PORTRAIT
    ws_cur.page_setup.paper_size = ws_cur.PAPERSIZE_TABLOID
    ws_cur.page_setup.fitToPage = True
    ws_cur.page_setup.fitToHeight = 1
    ws_cur.page_setup.fitToWidth = 1
    ws_cur.print_options.horizontalCentered = True
    ws_cur.add_print_title(6)
    ws_cur.page_margins = PageMargins(left=.5, right=.5, top=.5, bottom=.5,
                                      footer=.5)
    
    date_cur = datetime.strftime(datetime.today(), 
                          "%a {dt.month}/{dt.day}/%Y %I:%M %p".format(
                              dt=datetime.today()))
    
    ws_cur.oddFooter.right.text = date_cur
    ws_cur.oddFooter.right.size = 8
    ws_cur.oddFooter.right.font = 'Arial'
    
    
    #Freeze Panes
    c1 = ws_cur.cell(row=7, column=1)
    ws_cur.freeze_panes = c1            
            
    return wb_cur


if __name__ == "__main__":
    build_consolidated_p_and_l()
    retrieve_level_1()
    retrieve_level_2()
    level_1_income()
    level_1_expense()
    level_2_income()
    level_2_expense()    
    