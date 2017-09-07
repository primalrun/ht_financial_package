from openpyxl.workbook import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles import PatternFill
import sql_update
import sql_retrieve
import calendar
from excel_column_number import dict_col_converter as dcc
import pypyodbc
import setup1
import excel_formulas

#retrieve data
def retrieve1(dict_db):    
    sql1 = """
        create table #level_1 (
            period date,
            sales decimal(18, 6),
            adj_gross_margin decimal(18, 6),
            ebitda decimal(18,6),
            capex decimal(18, 6)
            )
        
        insert into #level_1
        select
            q2.period,
            q2.Sales,
            (q2.Sales - q2.Cost + q2.Adjustment_to_Margin) as Adjusted_Gross_Margin,
            (q2.Sales - q2.Cost + q2.Adjustment_to_Margin) - 
                (q2.Commissions + q2.Other_Personnel_Expense + q2.Other_Operating_Expense) + 
                (q2.Other_Income_Expense) 
                as EBITDA,
            q2.Capex
        from (
        
        select
            q1.period,
            sum(q1.Sales) as Sales,
            sum(q1.Cost) as Cost,
            sum(q1.Adjustment_to_Margin) as Adjustment_to_Margin,
            sum(q1.Commissions) as Commissions,
            sum(q1.Other_Personnel_Expense) as Other_Personnel_Expense,
            sum(q1.Other_Operating_Expense) as Other_Operating_Expense,
            sum(q1.Other_Income_Expense) as Other_Income_Expense,
            sum(q1.Capex) as Capex
        from (
        
        select
            b.period,
            -sum(case when r.level_1 = 'Sales' and r.company ='HT' then b.amount else 0 end) as Sales,
            sum(case when r.level_1 = 'Cost' and r.company ='HT' then b.amount else 0 end) as Cost,
            -sum(case when r.level_1 = 'Adjustment to Margin' and r.company ='HT' then b.amount else 0 end) as Adjustment_to_Margin,
            sum(case when r.level_1 = 'Commissions' and r.company ='HT' then b.amount else 0 end) as Commissions,
            sum(case when r.level_1 = 'Other Personnel Expense' and r.company ='HT' then b.amount else 0 end) as Other_Personnel_Expense,
            sum(case when r.level_1 = 'Other Operating Expense' and r.company ='HT' then b.amount else 0 end) as Other_Operating_Expense,
            -sum(case when r.level_1 = 'Other Income / (Expense)' and r.company ='HT' then b.amount else 0 end) as Other_Income_Expense,
            sum(case when r.level_1 = 'Capex' and r.company ='HT' then b.amount else 0 end) as Capex
        from Playground.[myop\jason.walker].gl_balance b
            inner join Playground.[myop\jason.walker].gl_account_reporting r
                on b.gl_account = r.gl_account
                and b.company = r.company
        group by
            b.period
        
        union all
        
        select
            b.period,
            -sum(case when r.level_1 = 'Sales' and r.company ='MYOP' then b.amount else 0 end) as Sales,
            sum(case when r.level_1 = 'Cost' and r.company ='MYOP' then b.amount else 0 end) as Cost,
            -sum(case when r.level_1 = 'Adjustment to Margin' and r.company ='MYOP' then b.amount else 0 end) as Adjustment_to_Margin,
            sum(case when r.level_1 = 'Commissions' and r.company ='MYOP' then b.amount else 0 end) as Commissions,
            sum(case when r.level_1 = 'Other Personnel Expense' and r.company ='MYOP' then b.amount else 0 end) as Other_Personnel_Expense,
            sum(case when r.level_1 = 'Other Operating Expense' and r.company ='MYOP' then b.amount else 0 end) as Other_Operating_Expense,
            -sum(case when r.level_1 = 'Other Income / (Expense)' and r.company ='MYOP' then b.amount else 0 end) as Other_Income_Expense,
            sum(case when r.level_1 = 'Capex' and r.company ='MYOP' then b.amount else 0 end) as Capex
        from Playground.[myop\jason.walker].gl_balance b
            inner join Playground.[myop\jason.walker].gl_account_reporting r
                on b.gl_account = r.gl_account
                and b.company = r.company
        group by
            b.period
        
        union all
        
        select
            b.period,
            -sum(case when r.level_1 = 'Sales' and r.company ='RAC' then b.amount else 0 end) as Sales,
            sum(case when r.level_1 = 'Cost' and r.company ='RAC' then b.amount else 0 end) as Cost,
            -sum(case when r.level_1 = 'Adjustment to Margin' and r.company ='RAC' then b.amount else 0 end) as Adjustment_to_Margin,
            sum(case when r.level_1 = 'Commissions' and r.company ='RAC' then b.amount else 0 end) as Commissions,
            sum(case when r.level_1 = 'Other Personnel Expense' and r.company ='RAC' then b.amount else 0 end) as Other_Personnel_Expense,
            sum(case when r.level_1 = 'Other Operating Expense' and r.company ='RAC' then b.amount else 0 end) as Other_Operating_Expense,
            -sum(case when r.level_1 = 'Other Income / (Expense)' and r.company ='RAC' then b.amount else 0 end) as Other_Income_Expense,
            sum(case when r.level_1 = 'Capex' and r.company ='RAC' then b.amount else 0 end) as Capex
        from Playground.[myop\jason.walker].gl_balance b
            inner join Playground.[myop\jason.walker].gl_account_reporting r
                on b.gl_account = r.gl_account
                and b.company = r.company
        group by
            b.period
        
        ) q1
        group by
            q1.period
        
        ) q2
        
        
        create table #level_2 (
            period date,
            excluded_expenses decimal(18, 6),
            deferred_partner_revenue decimal(18, 6),
            interest_expense decimal(18,6)
            )
        
        insert into #level_2
        select
            q1.period,
            sum(q1.excluded_expenses) as excluded_expenses,
            sum(q1.deferred_partner_revenue) as deferred_partner_revenue,
            sum(q1.interest_expense) as interest_expense
        from (
        
        select
            b.period,
            -sum(case when r.level_2 = 'Excluded Expenses' and r.company ='HT' then b.amount else 0 end) as excluded_expenses,
            -sum(case when r.level_2 = 'Deferred Partner Revenue' and r.company ='HT' then b.amount else 0 end) as deferred_partner_revenue,
            sum(case when r.level_2 = 'Interest Expense' and r.company ='HT' then b.amount else 0 end) as interest_expense
        from Playground.[myop\jason.walker].gl_balance b
            inner join Playground.[myop\jason.walker].gl_account_reporting r
                on b.gl_account = r.gl_account
                and b.company = r.company
        group by
            b.period
        
        union all
        
        select
            b.period,
            -sum(case when r.level_2 = 'Excluded Expenses' and r.company ='MYOP' then b.amount else 0 end) as excluded_expenses,
            -sum(case when r.level_2 = 'Deferred Partner Revenue' and r.company ='MYOP' then b.amount else 0 end) as deferred_partner_revenue,
            sum(case when r.level_2 = 'Interest Expense' and r.company ='MYOP' then b.amount else 0 end) as interest_expense
        from Playground.[myop\jason.walker].gl_balance b
            inner join Playground.[myop\jason.walker].gl_account_reporting r
                on b.gl_account = r.gl_account
                and b.company = r.company
        group by
            b.period
        
        union all
        
        select
            b.period,
            -sum(case when r.level_2 = 'Excluded Expenses' and r.company ='RAC' then b.amount else 0 end) as excluded_expenses,
            -sum(case when r.level_2 = 'Deferred Partner Revenue' and r.company ='RAC' then b.amount else 0 end) as deferred_partner_revenue,
            sum(case when r.level_2 = 'Interest Expense' and r.company ='RAC' then b.amount else 0 end) as interest_expense
        from Playground.[myop\jason.walker].gl_balance b
            inner join Playground.[myop\jason.walker].gl_account_reporting r
                on b.gl_account = r.gl_account
                and b.company = r.company
        group by
            b.period
        
        ) q1
        group by
            q1.period
        """
    sql2 = """ select
            l1.period,
            l1.sales,
            l1.adj_gross_margin,
            l1.ebitda,
            l2.excluded_expenses,
            l2.deferred_partner_revenue,
            l2.interest_expense,
            (l1.ebitda  - l2.excluded_expenses) as adj_ebitda,
            l1.capex - l1a.capex as capex
        from #level_1 l1
            left join #level_2 l2
                on l1.period = l2.period
            left join #level_1 l1a
                on l1.period = dateadd("m", +1, l1a.period)
    """
    
    connection_string = '''Driver={d1};
                        Server={s1};
                        Database={db1};
                        Trusted_Connection=Yes;
                    '''.format(d1=dict_db.get('driver'), 
                               s1=dict_db.get('server_2'), 
                               db1=dict_db.get('db_playground'))
                        
    connection = pypyodbc.connect(connection_string)    
    cur = connection.cursor()
    cur.execute(sql1)
    cur.commit()
    gl_data = cur.execute(sql2).fetchall()
    cur.close()
    connection.close    
    return gl_data

     
    

def excelupdate(Workbook, dict_db, date_end, month_actual, month_budget):
    #Get data from data warehouse
    gl_data = retrieve1(dict_db=dict_db)
    
    wb_new = Workbook
    ws_cur = wb_new.create_sheet('EBITDA Cash Analysis')
    
    col_actual_start = 2
    col_actual_end = col_actual_start + len(month_actual) - 1
    col_budget_start = col_actual_end + 1
    col_budget_end = col_budget_start + len(month_budget) - 1
    
    
    c1 = ws_cur.cell(row=1, column=1, value="Consolidated (HT/MYOP/RAC)")
    c1.alignment = Alignment(horizontal='center')
    c1.font = Font(bold='true')    
    c1 = ws_cur.cell(row=2, column=1, value="EBITDA Cash Analysis")
    c1.alignment = Alignment(horizontal='center')
    c1.font = Font(bold='true')
    str_period = ("For Period Ending {d1:%B} "
                 "{d1.day}, {d1.year}"
                 ).format(d1=date_end)    
    c1 = ws_cur.cell(row=3, column=1, value=str_period)
    c1.alignment = Alignment(horizontal='center')
    c1.font = Font(bold='true')
    row_date = 7
    
    ws_cur.merge_cells(start_row=1, end_row=1, start_column=1, 
                       end_column=14)
    ws_cur.merge_cells(start_row=2, end_row=2, start_column=1, 
                       end_column=14)
    ws_cur.merge_cells(start_row=3, end_row=3, start_column=1, 
                       end_column=14)        
    
    for c in range(2, 14):
        c1 = ws_cur.cell(row=row_date, column=c, 
                         value=calendar.month_name[c-1])
        c1.font = Font(underline='single', bold='true')
        c1.alignment = Alignment(horizontal='center')
        
    c1 = ws_cur.cell(row=row_date, column=14, 
                     value='Total ' + str(date_end.year))
    c1.font = Font(underline='single', bold='true')
    c1.alignment = Alignment(horizontal='center')
    
    col_last = ws_cur.max_column
        
    #Actual Header
    c1 = ws_cur.cell(row=row_date-1, column=col_actual_start, value='Actual')
    c1.font = Font(bold='true')
    c1.alignment =Alignment(horizontal='center')
    ws_cur.merge_cells(start_row=row_date-1, end_row=row_date-1, 
                       start_column=col_actual_start, 
                       end_column=col_actual_end)
    for c in range(col_actual_start, col_actual_end + 1):
        c1 = ws_cur.cell(row=row_date-1, column=c)
        c1.fill = PatternFill(start_color='b0e0e6', end_color='b0e0e6', fill_type='solid')

    #Budget Header
    c1 = ws_cur.cell(row=row_date-1, column=col_budget_start, value='Budget')
    c1.font = Font(bold='true')
    c1.alignment =Alignment(horizontal='center')
    ws_cur.merge_cells(start_row=row_date-1, end_row=row_date-1, 
                       start_column=col_budget_start, 
                       end_column=col_budget_end)
    for c in range(col_budget_start, col_budget_end + 1):
        c1 = ws_cur.cell(row=row_date-1, column=c)
        c1.fill = PatternFill(start_color='bdb76b', end_color='bdb76b', fill_type='solid')        
        
    
    r_next = row_date + 1
    
    #Sales
    c1 = ws_cur.cell(row=r_next, column=1, value='Sales')
    
    sales=[]
    for x in range(1, len(month_actual) + 1):
        sales.append(sum(r[1] for r in gl_data if r[0].month == x 
                          if r[0].year == date_end.year))    
    
    for c in range(2, 2 + len(month_actual)):
        c1 = ws_cur.cell(row=r_next, column=c, value=sales[c - 2])    
    
    adj_gross_margin = []
    for x in range(1, len(month_actual) + 1):
        adj_gross_margin.append(sum(r[2] for r in gl_data if r[0].month == x 
                          if r[0].year == date_end.year))
    
    r_next += 1
    c1 = ws_cur.cell(row=r_next, column=1, value='Adjusted Gross Margin')
    
    for c in range(2, 2 + len(month_actual)):
        c1 = ws_cur.cell(row=r_next, column=c, value=adj_gross_margin[c-2])
    
    for c in range(2, 15):
        for r in range(r_next-1, r_next+1):
            c1 = ws_cur.cell(row=r, column=c)
            c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
            
    #Adjusted Gross Margin %
    r_next += 1  
    c1 = ws_cur.cell(row=r_next, column=1, value='%')  
    for c in range(2, 15):
        formula1 = ("=if({col_letter}{row1}=0,0,"
                   "{col_letter}{row2}/{col_letter}{row1})" 
                   ).format(col_letter=dcc.get(c), 
                            row1=r_next - 2, row2=r_next-1)
        c1 = ws_cur.cell(row=r_next, column=c, value=formula1)
        c1.number_format = '0.0%'
    
    #Adjusted EBITDA
    adj_ebitda = []
    for x in range(1, len(month_actual) + 1):
        adj_ebitda.append(sum(r[7] for r in gl_data if r[0].month == x 
                          if r[0].year == date_end.year))
    
    r_next += 1
    c1 = ws_cur.cell(row=r_next, column=1, value='Adjusted EBITDA')
    
    for c in range(col_actual_start, col_actual_end + 1):
        c1 = ws_cur.cell(row=r_next, column=c, value=adj_ebitda[c-2])    
        
    for c in range(col_actual_start, col_last + 1):
        c1 = ws_cur.cell(row=r_next, column=c)
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
        
    #Adj EBITDA %
    r_next += 1  
    c1 = ws_cur.cell(row=r_next, column=1, value='%')  
    for c in range(2, 15):
        formula1 = ("=if({col_letter}{row1}=0,0,"
                   "{col_letter}{row2}/{col_letter}{row1})" 
                   ).format(col_letter=dcc.get(c), 
                            row1=r_next - 4, row2=r_next-1)
        c1 = ws_cur.cell(row=r_next, column=c, value=formula1)
        c1.number_format = '0.0%'
                
    #EBITDA Adjustment
    r_next += 2
    c1 = ws_cur.cell(row=r_next, column=1, value='EBITDA Adjustment')    
    r_next += 1
    c1 = ws_cur.cell(row=r_next, column=1, value='Adjusted EBITDA')
    c1.alignment = Alignment(indent=2)

    for c in range(col_actual_start, col_actual_end + 1):
        c1 = ws_cur.cell(row=r_next, column=c, value=adj_ebitda[c-2])
        
    def_revenue = []
    for x in range(1, len(month_actual) + 1):
        def_revenue.append(sum(r[5] for r in gl_data if r[0].month == x 
                          if r[0].year == date_end.year))    
    r_next += 1
    c1 = ws_cur.cell(row=r_next, column=1, value='Deferred Revenue')
    c1.alignment = Alignment(indent=2)
    
    for c in range(col_actual_start, col_actual_end + 1):
        c1 = ws_cur.cell(row=r_next, column=c, value=def_revenue[c-2])
        
    excl_expense = []
    for x in range(1, len(month_actual) + 1):
        excl_expense.append(sum(r[4] for r in gl_data if r[0].month == x 
                          if r[0].year == date_end.year))    
    r_next += 1
    c1 = ws_cur.cell(row=r_next, column=1, value='Excluded Expenses')
    c1.alignment = Alignment(indent=2)
    
    for c in range(col_actual_start, col_actual_end + 1):
        c1 = ws_cur.cell(row=r_next, column=c, value=excl_expense[c-2])
        
    r_next += 1
    c1 = ws_cur.cell(row=r_next, column=1, value='Cash EBITDA')    
    for c in range(col_actual_start, col_last + 1):
        formula1 = "=sum({col_letter}{row1}:{col_letter}{row2})".format(
            col_letter=dcc.get(c), row1=r_next-3, row2=r_next-1)        
        c1 = ws_cur.cell(row=r_next, column=c, value=formula1)
    
    #Cash Needs
    r_next += 2
    c1 = ws_cur.cell(row=r_next, column=1, value='Cash Needs')
    r_next += 1
    c1 = ws_cur.cell(row=r_next, column=1, value='Cash Interest (1)')    
    c1.alignment = Alignment(indent=2)
    def_finance_fees = setup1.def_finance_fees()
    interest_expense = []
    
    for x in range(1, len(month_actual) + 1):
        interest_expense.append(sum(r[6] - def_finance_fees for r in gl_data 
                                    if r[0].month == x 
                                    if r[0].year == date_end.year))
    
    for c in range(col_actual_start, col_actual_end + 1):
        c1 = ws_cur.cell(row=r_next, column=c, value=interest_expense[c-2])
        
    r_next += 1
    c1 = ws_cur.cell(row=r_next, column=1, value='Principal Payments (2)')        
    c1.alignment = Alignment(indent=2)
    
    for c in range(col_actual_start, col_actual_end + 1):
        c1 = ws_cur.cell(row=r_next, column=c, 
                         value=setup1.principal_payments())    

    r_next += 1
    c1 = ws_cur.cell(row=r_next, column=1, value='CAPEX')
    c1.alignment = Alignment(indent=2)
    capex = []
    
    for x in range(1, len(month_actual) + 1):
        capex.append(sum(r[8] for r in gl_data 
                                    if r[0].month == x 
                                    if r[0].year == date_end.year))
            
    for c in range(col_actual_start, col_actual_end + 1):
        c1 = ws_cur.cell(row=r_next, column=c, value=capex[c-2])            

    r_next += 1
    c1 = ws_cur.cell(row=r_next, column=1, value='Total Cash Needs')
        
    for c in range(col_actual_start, col_last + 1):
        formula1 = "=sum({col_letter}{row1}:{col_letter}{row2})".format(
            col_letter=dcc.get(c), row1=r_next-3, row2=r_next-1)        
        c1 = ws_cur.cell(row=r_next, column=c, value=formula1)
        
    #Free Cash Flow
    r_next += 2
    c1 = ws_cur.cell(row=r_next, column=1, 
                     value='Free Cash Flow (EBITDA Coverage')
    
    for c in range(col_actual_start, col_last + 1):
        formula1 = "={col_letter}{row1}-{col_letter}{row2}".format(
            col_letter=dcc.get(c), row1=r_next-8, row2=r_next-2)        
        c1 = ws_cur.cell(row=r_next, column=c, value=formula1)        
        
    #Budgeted FCF
    r_next += 1
    c1 = ws_cur.cell(row=r_next, column=1, 
                     value='Budged FCF')        
    
    #Variance
    r_next += 1
    c1 = ws_cur.cell(row=r_next, column=1, 
                     value='Variance')        
    
    for c in range(col_actual_start, col_last + 1):
        formula1 = "={col_letter}{row1}-{col_letter}{row2}".format(
            col_letter=dcc.get(c), row1=r_next-2, row2=r_next-1)        
        c1 = ws_cur.cell(row=r_next, column=c, value=formula1)    
    
    
    #Total Year
    sum_rows = [8, 9, 11, 15, 16, 17, 21, 22, 23, 27]    
    
    for r in range(0, len(sum_rows)):
        ws_cur.cell(row=sum_rows[r], column=col_last, 
                    value =excel_formulas.sum_row_1(sum_rows[r], 
                                                    2, 
                                                    col_last - 1))
    
    
    #Format Cells
    row_start = row_date + 6
    row_last = ws_cur.max_row
    
    
    for c in range(2, col_last + 1):
        for r in range(row_start, row_last + 1):
            c1 = ws_cur.cell(row=r, column=c)
            c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    
    #Footnotes
    r_next = row_last + 3
    str_note = ("1) Deferred Financing Fees are deducted from interest expense "
                "since it is non cash.")
    c1 = ws_cur.cell(row=r_next, column=1, value=str_note)
    
    
    #Column Width
    ws_cur.column_dimensions[dcc.get(1)].width = 35
    
    for c in range(2, 15):
        ws_cur.column_dimensions[dcc.get(c)].width = 13
    
    #Page Setup
    ws_cur.page_setup.orientation = ws_cur.ORIENTATION_LANDSCAPE
    ws_cur.page_setup.paper_size = ws_cur.PAPERSIZE_TABLOID
    ws_cur.page_setup.fitToPage = True
    ws_cur.page_setup.fitToHeight = False
    ws_cur.page_setup.fitToWidth = 1
    ws_cur.print_options.horizontalCentered = True
    ws_cur.add_print_title(7)
    ws_cur.page_margins = PageMargins(left=.5, right=.5, top=.5, bottom=.5)
    
    #Freeze Panes
    c1 = ws_cur.cell(row=8, column=1)
    ws_cur.freeze_panes = c1    
                         
    return wb_new

if __name__ == "__main__":
    excelupdate()
    retrieve1()


