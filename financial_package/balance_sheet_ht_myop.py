import sql_retrieve
from dateutil import relativedelta
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.worksheet.page import PageMargins
from datetime import datetime
from openpyxl.styles import Border
from openpyxl.styles import Side

def retrieve_bs_data(dict_db):
    sql = """
    select
        glr.level_1,
        glr.level_2,
        glb.period,
        sum(glb.amount) as amount
    from Playground.[myop\jason.walker].gl_balance_bs glb
        inner join Playground.[myop\jason.walker].gl_account_reporting glr
            on glb.company = glr.company
            and glb.gl_account = glr.gl_account
    where
        glr.financial_statement = 'BS'
        and glr.level_1 is not null
        and glr.level_2 is not null
        and glb.company in ('HT', 'MYOP')
    group by
        glr.level_1,
        glr.level_2,
        glb.period
        
        union all
        
        select
            'Net Income',
            'Net Income',
            glb.period,
            sum(glb.amount) as amount
        from Playground.[myop\jason.walker].gl_balance_bs glb
        where
            glb.gl_account = '99999'
            and glb.company in ('HT', 'MYOP')
        group by
            glb.period
                    
        union all
        
        select
            'Net Income YTD',
            'Net Income YTD',
            glb.period,
            sum(glb.amount) as amount
        from Playground.[myop\jason.walker].gl_balance_bs glb
        where
            glb.gl_account = '99999Z'
            and glb.company in ('HT', 'MYOP')
        group by
            glb.period                    
    """
    
    gl_data = sql_retrieve.getdata1(a_driver=dict_db.get('driver'), 
                                    a_server=dict_db.get('server_2'), 
                                    a_db=dict_db.get('db_playground'), 
                                    a_sql=sql)

    return gl_data



def build_consolidated_bs(wb, dict_db, date_end):
    wb_cur = wb
    date_prior_ye = (date_end
                      + relativedelta.relativedelta(months=-date_end.month))

    
    gl_data = retrieve_bs_data(dict_db)
    assets = sorted(set([x[1] for x in gl_data if x[0] == 'Asset']))
    liabilities = sorted(set([x[1] for x in gl_data if x[0] == 'Liability']))
    equities = sorted(set([x[1] for x in gl_data if x[0] == 'Equity']))
    
    ws_cur = wb_cur.create_sheet('BS_HT_MYOP')
    c1=ws_cur.cell(row=1, 
                column=1, 
                value='Consolidated Balance Sheet - (HT/MYOP)')
    c1.font = Font(bold='true')
    c1.alignment = Alignment(horizontal='center')
    
    ws_cur.merge_cells(start_row=1, 
                       end_row=1, 
                       start_column=1,
                       end_column=3)
    
    ws_cur.cell(row=3, column=2, 
                value=datetime.strftime(date_end, '%B %d, %Y'))
    ws_cur.cell(row=3, column=3, 
                value=datetime.strftime(date_prior_ye, '%B %d, %Y'))
    
    for c in range(2, 4):
        c1 = ws_cur.cell(row=3, column=c)
        c1.fill = PatternFill(start_color='1e90ff',
                              end_color='1e90ff',
                              fill_type='solid')
        c1.font = Font(bold='true', color='f8f8ff')
        c1.alignment = Alignment(horizontal='center')
    
    r_next = 4
    c1=ws_cur.cell(row=r_next, column=1, value='ASSETS:')
    c1.font = Font(bold='true')
    
    r_next += 1
    
    #Border Style
    top_border = Border(top=Side(style='thin'))
    
    #Assets
    r = 0
    
    for a in assets:
        ws_cur.cell(row=r_next + r, column=1, value=a)
        cur_period_bal = sum([x[3] for x in gl_data 
                          if x[0] == 'Asset' 
                          and x[1] == a 
                          and x[2] == date_end])
        ws_cur.cell(row=r_next + r, column=2, value=cur_period_bal)
        py_end_bal = sum([x[3] for x in gl_data 
                          if x[0] == 'Asset' 
                          and x[1] == a 
                          and x[2] == date_prior_ye])
        ws_cur.cell(row=r_next + r, column=3, value=py_end_bal)        
        r += 1    
           
    r_next = r_next + r
    r_asset = r_next
    
    c1 = ws_cur.cell(row=r_next, column=1, value='Total Assets')
    c1.font = Font(bold='true')
    
    formula1 = '=sum(B{r1}:B{r2})'.format(
        r1=r_next - len(assets), r2=r_next - 1)
    c1 = ws_cur.cell(row=r_next, column=2, value=formula1)    
    c1.font = Font(bold='true')
    c1.border = top_border
    
    formula1 = '=sum(C{r1}:C{r2})'.format(
        r1=r_next - len(assets), r2=r_next - 1)
    c1 = ws_cur.cell(row=r_next, column=3, value=formula1)    
    c1.font = Font(bold='true')
    c1.border = top_border
    
    #Liabilities
    r_next += 2
    c1 = ws_cur.cell(row=r_next, column=1, value='LIABILITIES AND EQUITY:')
    c1.font = Font(bold='true')
    
    r_next += 1
    r = 0
    
    for liab in liabilities:
        ws_cur.cell(row=r_next + r, column=1, value=liab)
        cur_period_bal = -sum([x[3] for x in gl_data 
                          if x[0] == 'Liability' 
                          and x[1] == liab 
                          and x[2] == date_end])
        ws_cur.cell(row=r_next + r, column=2, value=cur_period_bal)
        py_end_bal = -sum([x[3] for x in gl_data 
                          if x[0] == 'Liability' 
                          and x[1] == liab 
                          and x[2] == date_prior_ye])
        ws_cur.cell(row=r_next + r, column=3, value=py_end_bal)        
        r += 1    
           
    r_next = r_next + r
    r_liability = r_next
    c1 = ws_cur.cell(row=r_next, column=1, value='Total Liabilities')
    c1.font = Font(bold='true')
    
    formula1 = '=sum(B{r1}:B{r2})'.format(
        r1=r_next - len(liabilities), r2=r_next - 1)
    c1 = ws_cur.cell(row=r_next, column=2, value=formula1)    
    c1.font = Font(bold='true')
    c1.border = top_border
    
    formula1 = '=sum(C{r1}:C{r2})'.format(
        r1=r_next - len(liabilities), r2=r_next - 1)
    c1 = ws_cur.cell(row=r_next, column=3, value=formula1)    
    c1.font = Font(bold='true')
    c1.border = top_border
    
    #Equity
    r_next += 2
    r = 0
    
    for e in equities:
        ws_cur.cell(row=r_next + r, column=1, value=e)
        cur_period_bal = -sum([x[3] for x in gl_data 
                          if x[0] == 'Equity' 
                          and x[1] == e 
                          and x[2] == date_end])
        ws_cur.cell(row=r_next + r, column=2, value=cur_period_bal)
        py_end_bal = -sum([x[3] for x in gl_data 
                          if x[0] == 'Equity' 
                          and x[1] == e 
                          and x[2] == date_prior_ye])
        ws_cur.cell(row=r_next + r, column=3, value=py_end_bal)        
        r += 1    
    
    #Net Income
    r_next = r_next + r
    r_net_income = r_next
    r_retained_earning = r_net_income - 1
    ws_cur.cell(row=r_next, column=1, value='Net Income')
    cur_period_bal = sum([x[3] for x in gl_data 
                      if x[0] == 'Net Income' 
                      and x[1] == 'Net Income' 
                      and x[2] == date_end])    
    ws_cur.cell(row=r_next, column=2, value=cur_period_bal)
    py_end_bal = sum([x[3] for x in gl_data 
                      if x[0] == 'Net Income' 
                      and x[1] == 'Net Income' 
                      and x[2] == date_prior_ye])    
    ws_cur.cell(row=r_next, column=3, value=py_end_bal)
    
    r_next += 1
    r_equity = r_next
    c1 = ws_cur.cell(row=r_next, column=1, value='Total Equity')
    c1.font = Font(bold='true')
    
    formula1 = '=sum(B{r1}:B{r2})'.format(
        r1=r_next - len(equities) - 1, r2=r_next - 1)
    c1 = ws_cur.cell(row=r_next, column=2, value=formula1)    
    c1.font = Font(bold='true')
    c1.border = top_border
    
    formula1 = '=sum(C{r1}:C{r2})'.format(
        r1=r_next - len(equities) - 1, r2=r_next - 1)
    c1 = ws_cur.cell(row=r_next, column=3, value=formula1)    
    c1.font = Font(bold='true')
    c1.border = top_border
    
    #Total Liabilities and Equity
    r_next += 2
    c1=ws_cur.cell(row=r_next, column=1, value='Total Liabilities and Equity')
    c1.font = Font(bold='true')
    
    formula1 = '=B{r1}+B{r2}'.format(
        r1=r_liability, r2=r_equity)
    c1 = ws_cur.cell(row=r_next, column=2, value=formula1)
    c1.font = Font(bold='true')
    
    formula1 = '=C{r1}+C{r2}'.format(
        r1=r_liability, r2=r_equity)
    c1 = ws_cur.cell(row=r_next, column=3, value=formula1)
    c1.font = Font(bold='true')
        
    r_next += 2
    c1 = ws_cur.cell(row=r_next, column=1, value='Accounting Equation')
    c1.font = Font(bold='true')
    
    formula1='=B{r1}-B{r2}'.format(
        r1=r_asset, r2=r_next - 2)
    c1 = ws_cur.cell(row=r_next, column=2, value=formula1)
    c1.font = Font(bold='true')
    
    formula1='=C{r1}-C{r2}'.format(
        r1=r_asset, r2=r_next - 2)
    c1 = ws_cur.cell(row=r_next, column=3, value=formula1)
    c1.font = Font(bold='true')    
    
    
    
    
    #Reclass non YTD Net Income to Retained Earnings---------------------------    
    net_income_cur = ws_cur.cell(row=r_net_income, column=2).value
    net_income_cur_ytd = sum([x[3] for x in gl_data 
                      if x[0] == 'Net Income YTD' 
                      and x[1] == 'Net Income YTD' 
                      and x[2] == date_end])    
    net_income_remainder = net_income_cur - net_income_cur_ytd
    retained_earning_cur = ws_cur.cell(row=r_retained_earning, column=2).value
    retained_earning_new = retained_earning_cur + net_income_remainder
    ws_cur.cell(row=r_retained_earning, column=2, value=retained_earning_new)
    ws_cur.cell(row=r_net_income, column=2, value=net_income_cur_ytd)
    
    net_income_pye = ws_cur.cell(row=r_net_income, column=3).value
    net_income_py_ytd = sum([x[3] for x in gl_data 
                      if x[0] == 'Net Income YTD' 
                      and x[1] == 'Net Income YTD' 
                      and x[2] == date_prior_ye])    
    net_income_remainder = net_income_pye - net_income_py_ytd
    retained_earning_pye = ws_cur.cell(row=r_retained_earning, column=3).value
    retained_earning_new = retained_earning_pye + net_income_remainder
    ws_cur.cell(row=r_retained_earning, column=3, value=retained_earning_new)
    ws_cur.cell(row=r_net_income, column=3, value=net_income_py_ytd)
    #-------------------------------------------------------------------------
    
    
    
    
    
    
    #Format Cells
    format_number = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    
    for r in range(5, ws_cur.max_row + 1):
        for c in range(2, 4):
            ws_cur.cell(row=r, column=c).number_format = format_number
    
      
    #Column widths
    ws_cur.column_dimensions['A'].width = 44
    ws_cur.column_dimensions['B'].width = 20
    ws_cur.column_dimensions['C'].width = 20    
    
    #Page Setup
    ws_cur.page_setup.orientation = ws_cur.ORIENTATION_PORTRAIT
    ws_cur.page_setup.paper_size = ws_cur.PAPERSIZE_TABLOID
    ws_cur.page_setup.fitToPage = True
    ws_cur.page_setup.fitToHeight = False
    ws_cur.page_setup.fitToWidth = 1
    ws_cur.print_options.horizontalCentered = True
    ws_cur.add_print_title(3)
    ws_cur.page_margins = PageMargins(left=.5, right=.5, top=.5, bottom=.5)
    
    #Freeze Panes
    c1 = ws_cur.cell(row=4, column=1)
    ws_cur.freeze_panes = c1            
    
    
        
    return wb_cur    
    


if __name__ == "__main__":
    retrieve_bs_data()
    build_consolidated_bs()