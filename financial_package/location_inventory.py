from openpyxl import load_workbook
from os import listdir
import sys
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from excel_column_number import dict_col_converter as dcc
from openpyxl.worksheet.page import PageMargins
from tkinter import Tk
from tkinter import messagebox

def location_inventory_update(wb):
    wb_cur = wb
    ws_cur = wb.create_sheet('Location Inventory')
    file_path = "Z:/Accounting/Accounting/Financial Package/"
    dir_list = listdir(file_path)
    inv_file = [x for x in dir_list if 'Inventory By Location' in x]
    
    
    if len(inv_file) > 1:
        print('Multiple Inventory By Location Files, Process Cancelled!')
        sys.exit()
    
    msg_text = "Use Inventory File {f1}?".format(f1=inv_file)
    root = Tk()
    root.withdraw()
    answer_loc_file = messagebox.askquestion('Location File', msg_text)
    
    if answer_loc_file == 'no':
        print('Wrong Location File in Package Folder, Process Cancelled!')
        err_index = 1
        cont_return = [wb_cur, err_index]
        sys.exit()
        
    wb_source = load_workbook(file_path + inv_file[0], read_only=True)
    
    #Last Worksheet
    ws_source = wb_source.worksheets[len(wb_source.worksheets) - 1]    
    
    #Last data row    
    r = 1
    
    while True:
        if ws_source.cell(row=r, column=2).value == 'TOTAL INVENTORY':
            break
        if r > ws_source.max_row:        
            break
        r += 1
    
    if r > ws_source.max_row:
        print(("Location Inventory: Can't find TOTAL INVENTORY record, "
               "Process Cancelled"))
        err_index = 1
        cont_return = [wb_cur, err_index]        
        sys.exit()    
    
    row_end = r - 1
    
    #Header data row    
    r = 1
    
    while True:
        if ws_source.cell(row=r, column=2).value == 'Location':
            break
        if r > ws_source.max_row:        
            break
        r += 1
    
    if r > ws_source.max_row:
        print(("Location Inventory: Can't find Location column header record, "
               "Process Cancelled"))
        err_index = 1
        cont_return = [wb_cur, err_index]        
        sys.exit()    
    
    row_header = r
    
    titles = []
    titles.append(ws_source.cell(row=1, column=1).value)
    titles.append(ws_source.cell(row=2, column=1).value)
    titles.append(ws_source.cell(row=3, column=1).value)
    
    header1 = [ws_source.cell(row=row_header,
                              column=x).value for x in range(1, 7)]
    
    #flat list of records
    records = []
    
    for r in range(row_header + 1, row_end + 1):
        for c in range(1, 5):
            records.append(ws_source.cell(row=r, column=c).value)
    
    #row list of records
    records2 = []
    
    for x in range(0, len(records), 4):
        records2.append(records[x:x+4])
        
    for x in range(0, len(records2)):        
        if records2[x][0] == None:            
            records2[x][0] = records2[x - 1][0]
        if records2[x][0] == 'Managed':            
            records2[x][0] = 'Vendor Managed'            
    
    
    #remove blanks
    records3 = [x for x in records2 if float(x[2] or 0) + float(x[3] or 0) != 0]        
        
    #Write to Financial Package
    c1 = ws_cur.cell(row=1, column=1, value=titles[0])
    c1.font = Font(bold='true')
    c1.alignment = Alignment(horizontal='center')
        
    c1 = ws_cur.cell(row=2, column=1, value=titles[1])
    c1.font = Font(bold='true')
    c1.alignment = Alignment(horizontal='center')
        
    c1 = ws_cur.cell(row=3, column=1, value=titles[2])
    c1.font = Font(bold='true')
    c1.alignment = Alignment(horizontal='center')    

    ws_cur.merge_cells(start_row=1, end_row=1,
                       start_column=1, end_column=6)

    
    ws_cur.merge_cells(start_row=2, end_row=2,
                       start_column=1, end_column=6)    

    ws_cur.merge_cells(start_row=3, end_row=3,
                       start_column=1, end_column=6)
    
    #Header
    r_next = 5
    
    for c in range(1, 7):
        c1 = ws_cur.cell(row=r_next, column=c, value=header1[c-1])
        c1.fill = PatternFill(start_color='cdc9c9', end_color='cdc9c9', 
                              fill_type='solid')    
    
    #Records
    r_next += 1    
    c_next = 1
    
    for r in range(0, len(records3)):
        for e in range(0, len(records3[r])):            
            ws_cur.cell(row=r_next, column=c_next, value=records3[r][e])
            c_next += 1
        c_next = 1    
        r_next += 1
    
    r_next = 6
    
    for r in range(r_next, ws_cur.max_row + 1):
        formula1 = "={col_1}{row_cur}-{col_2}{row_cur}".format(
            row_cur=r, col_1=dcc.get(4), col_2=dcc.get(3))
        ws_cur.cell(row=r, column=5, value=formula1)
        
    for r in range(r_next, ws_cur.max_row + 1):
        formula1 = ("=if({col_1}{row_cur}=0,0,"
                    "{col_2}{row_cur}/{col_1}{row_cur})").format(
            row_cur=r, col_1=dcc.get(3), col_2=dcc.get(5))
        ws_cur.cell(row=r, column=6, value=formula1)    
    
    
    #Count region records
    regions = sorted(set([x[0] for x in records3]))
    region_count = {}
    
    for r in regions:
        region_count[r] = [x[0] for x in records3 if x[0] == r].count(r)
    
    
    #Totals
    r_next = ws_cur.max_row + 2
    ws_cur.cell(row=r_next, column=1, value='Total')
    inv_pm = sum([x[2] for x in records3])
    ws_cur.cell(row=r_next, column=3, value=inv_pm)
    inv_cm = sum([x[3] for x in records3])
    ws_cur.cell(row=r_next, column=4, value=inv_cm)    
    
    formula1 = "={col_1}{row_cur}-{col_2}{row_cur}".format(
        row_cur=r_next, col_1=dcc.get(4), col_2=dcc.get(3))
    ws_cur.cell(row=r_next, column=5, value=formula1)
    
    formula1 = ("=if({col_1}{row_cur}=0,0,"
                "{col_2}{row_cur}/{col_1}{row_cur})").format(
        row_cur=r_next, col_1=dcc.get(3), col_2=dcc.get(5))
    ws_cur.cell(row=r_next, column=6, value=formula1)        
    
    for c in range(1, 7):
        ws_cur.cell(row=r_next, column=c).font = Font(bold=True)
            
    #Region Totals
    r_next = ws_cur.max_row + 2
    regions_supplies = [x for x in regions if x != 'Vendor Managed']
    
    for r in range(0, len(regions_supplies)):
        ws_cur.cell(row=r_next, column=1, value=regions_supplies[r])
        inv_pm = sum([x[2] for x in records3
                                if x[0] == regions_supplies[r]])
        ws_cur.cell(row=r_next, column=3, value=inv_pm)
        inv_cm = sum([x[3] for x in records3
                                if x[0] == regions_supplies[r]])
        ws_cur.cell(row=r_next, column=4, value=inv_cm)        
        
        formula1 = "={col_1}{row_cur}-{col_2}{row_cur}".format(
            row_cur=r_next, col_1=dcc.get(4), col_2=dcc.get(3))
        ws_cur.cell(row=r_next, column=5, value=formula1)
        
        formula1 = ("=if({col_1}{row_cur}=0,0,"
                    "{col_2}{row_cur}/{col_1}{row_cur})").format(
            row_cur=r_next, col_1=dcc.get(3), col_2=dcc.get(5))
        ws_cur.cell(row=r_next, column=6, value=formula1)        
        
        
        r_next += 1
        
    r_next += 1
    ws_cur.cell(row=r_next, column=1, value='Total Supplies Inventory')    
    inv_pm_supplies = sum([x[2] for x in records3
                            if x[0] != 'Vendor Managed'])
    ws_cur.cell(row=r_next, column=3, value=inv_pm_supplies)
    inv_cm_supplies = sum([x[3] for x in records3
                            if x[0] != 'Vendor Managed'])
    ws_cur.cell(row=r_next, column=4, value=inv_cm_supplies)
    
    formula1 = "={col_1}{row_cur}-{col_2}{row_cur}".format(
        row_cur=r_next, col_1=dcc.get(4), col_2=dcc.get(3))
    ws_cur.cell(row=r_next, column=5, value=formula1)
    
    formula1 = ("=if({col_1}{row_cur}=0,0,"
                "{col_2}{row_cur}/{col_1}{row_cur})").format(
        row_cur=r_next, col_1=dcc.get(3), col_2=dcc.get(5))
    ws_cur.cell(row=r_next, column=6, value=formula1)        
        
    for c in range(1, 7):
        ws_cur.cell(row=r_next, column=c).font = Font(bold=True)        
        
    #Vendor Managed Total
    r_next += 2
    ws_cur.cell(row=r_next, column=1, value='Vendor Managed')    
    inv_pm_vendor = sum([x[2] for x in records3
                            if x[0] == 'Vendor Managed'])
    ws_cur.cell(row=r_next, column=3, value=inv_pm_vendor)
    inv_cm_vendor = sum([x[3] for x in records3
                            if x[0] == 'Vendor Managed'])
    ws_cur.cell(row=r_next, column=4, value=inv_cm_vendor)
    
    formula1 = "={col_1}{row_cur}-{col_2}{row_cur}".format(
        row_cur=r_next, col_1=dcc.get(4), col_2=dcc.get(3))
    ws_cur.cell(row=r_next, column=5, value=formula1)
    
    formula1 = ("=if({col_1}{row_cur}=0,0,"
                "{col_2}{row_cur}/{col_1}{row_cur})").format(
        row_cur=r_next, col_1=dcc.get(3), col_2=dcc.get(5))
    ws_cur.cell(row=r_next, column=6, value=formula1)            
    
    for c in range(1, 7):
        ws_cur.cell(row=r_next, column=c).font = Font(bold=True)    
    
    #Format
    format_number = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    for r in range(6, ws_cur.max_row + 1):
        for c in range(3, 6):
            ws_cur.cell(row=r, column=c).number_format = format_number
    
    for r in range(6, ws_cur.max_row + 1):
        for c in range(6, 7):
            ws_cur.cell(row=r, column=c).number_format = '0.0%'
    
    for c in range(1, 7):
        ws_cur.cell(row=5, column=c).alignment = Alignment(horizontal='center')
        ws_cur.cell(row=5, column=c).font = Font(bold=True)
        
    #Page Setup
    ws_cur.column_dimensions[dcc.get(1)].width = 24
    ws_cur.column_dimensions[dcc.get(2)].width = 33
    ws_cur.column_dimensions[dcc.get(3)].width = 18
    ws_cur.column_dimensions[dcc.get(4)].width = 18
    ws_cur.column_dimensions[dcc.get(5)].width = 18
    ws_cur.column_dimensions[dcc.get(6)].width = 13
    
    
    ws_cur.page_setup.orientation = ws_cur.ORIENTATION_PORTRAIT
    ws_cur.page_setup.paper_size = ws_cur.PAPERSIZE_TABLOID
    ws_cur.page_setup.fitToPage = True
    ws_cur.page_setup.fitToHeight = 1
    ws_cur.page_setup.fitToWidth = 1
    ws_cur.print_options.horizontalCentered = True
    ws_cur.add_print_title(5)
    ws_cur.page_margins = PageMargins(left=.5, right=.5, top=.5, bottom=.5,
                                      footer=.5)
    
    #Freeze Panes
    c1 = ws_cur.cell(row=6, column=1)
    ws_cur.freeze_panes = c1                
    
    
    
    cont_return = [wb_cur]    
    return   cont_return











if __name__ == "__main__":
    location_inventory_update()
    