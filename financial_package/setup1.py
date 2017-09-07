import sys
import os
from datetime import date
from time import mktime
from time import strptime
from tkinter import messagebox
from tkinter import Tk
from calendar import monthrange
from dateutil.relativedelta import relativedelta
from openpyxl.drawing.image import Image
from excel_column_number import dict_col_converter as dcc
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from datetime import datetime

def setup_f1():
    #Dates
    date_today = date.today()
    
    if date_today.month == 1:
        prior_month = 12
        prior_month_year = date_today.year - 1
    else:
        prior_month = date_today.month - 1
        prior_month_year = date_today.year
        
    date_start = date(prior_month_year, prior_month, 1) 
    
    msg_text = 'Use close period starting {d1}'.format(
        d1=date_start)
    
    #Ask user to validate close period
    root = Tk()
    root.withdraw()
    answer_date = messagebox.askquestion('Date Validation', msg_text)
    
    if answer_date == 'no':
        date_input = input('Type close period start date m/d/yyyy')    
        try:
            valid_date = strptime(date_input, '%m/%d/%Y')       
            date_start = date.fromtimestamp(mktime(valid_date))
        except:
            print('Invalid Date, Process Cancelled')
            sys.exit()
    
    date_end = date(date_start.year, 
                date_start.month, 
                monthrange(date_start.year, date_start.month)[1])
        
    #Directory check
    path_str = "c:/temp/"
    if not os.path.exists(path_str):
        os.makedirs(path_str)
        
    cont_1 = [date_start, path_str, date_end, date_ytd_start]
    return cont_1    

def db_dictionary():
    dict_db = {}
    dict_db['driver'] = 'SQL Server Native Client 11.0'
    dict_db['server_3'] = 'TNDCSQL03'
    dict_db['server_2'] = 'TNDCSQL02'
    dict_db['db_gl'] = 'NAVREP'
    dict_db['db_playground'] = 'Playground'
    dict_db['table_schema_1'] = 'dbo'
    dict_db['table_schema_2'] = '[myop\jason.walker]'
    dict_db['table_gl_ht'] = '[Hi Touch$G_L Entry]'
    dict_db['table_gl_myop'] = '[MYOP Touch$G_L Entry]'
    dict_db['table_gl_rac'] = '[Rentacrate Touch$G_L Entry]'
    dict_db['table_gl_reporting'] = 'gl_account_reporting'
    dict_db['table_gl_balance'] = 'gl_balance'     
    return dict_db

def def_finance_fees():
    return 63623
    
def principal_payments():
    return 0

def date_ytd_start(date_start):
    date_ytd_start = date_start + relativedelta(months=-date_start.month + 1)
    return date_ytd_start

def date_trail_12_start(date_start):
    date_trail_12_start = date_start + relativedelta(months=-11)
    return date_trail_12_start

def get_cover_page(wb, template_directory, date_end):
    wb_cur = wb
    ws_cur = wb_cur.create_sheet('Cover')
    
    ht_logo = Image(template_directory + 'ht_logo.png')
    ht_logo.anchor(ws_cur.cell(row=7, column=3), 'absolute')
    ws_cur.add_image(ht_logo)

    companies_logo = Image(template_directory + 'companies_logo.png') 
    companies_logo.anchor(ws_cur.cell(row=31, column=1), 'absolute')
    ws_cur.add_image(companies_logo)        
        
    for c in range(1, 13):
        ws_cur.column_dimensions[(dcc.get(c))].width = 8.5
    
    c1 = ws_cur.cell(row=17, column=1, value='Consolidated Financial Results')
    c1.alignment = Alignment(horizontal='center')
    c1.font = Font(bold='true', size=22)
    ws_cur.merge_cells(start_row=17, end_row=17, start_column=1, 
                       end_column=12)
    
    c1 = ws_cur.cell(row=22, column=1, 
                     value=datetime.strftime(date_end, '%B %Y'))
    c1.alignment = Alignment(horizontal='center')
    c1.font = Font(bold='true', size=22)
    ws_cur.merge_cells(start_row=22, end_row=22, start_column=1, 
                       end_column=12)
    
        
    ws_cur.page_setup.orientation = ws_cur.ORIENTATION_PORTRAIT
    ws_cur.page_setup.paper_size = ws_cur.PAPERSIZE_TABLOID
    ws_cur.page_setup.fitToPage = True
    ws_cur.page_setup.fitToHeight = 1
    ws_cur.page_setup.fitToWidth = 1
    ws_cur.print_options.horizontalCentered = True    
    ws_cur.page_margins = PageMargins(left=.5, right=.5, top=.5, bottom=.5,
                                      footer=.5)    
    
    return wb_cur


    
if __name__ == "__main__":
    setup_f1()
    def_finance_fees()
    principal_payments()
    date_ytd_start()
    date_trail_12_start()
    get_cover_page()
    db_dictionary()
        

