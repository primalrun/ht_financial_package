import sql_retrieve
from openpyxl.styles import Font
from datetime import datetime
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border
from openpyxl.styles.borders import Side
from excel_column_number import dict_col_converter as dcc
from openpyxl.worksheet.page import PageMargins

def get_rebate_data(date_end, dict_db):
    sql = """
    declare @dateend date = '{d1}'
    declare @datestart date = dateadd("m", -11, @dateend)
    set @datestart = dateadd("d", -datepart("d", @datestart) + 1, @datestart)
    
    select
        glb.period,
        case when (glb.company = 'HT' and glb.gl_account = '45650') or (glb.company = 'MYOP' and glb.gl_account = '52000') then 'Pricing'
            when glb.company = 'MYOP' and glb.gl_account = '52100' then 'Wholesalers'
            when (glb.company = 'HT' and glb.gl_account = '45660') or (glb.company = 'MYOP' and glb.gl_account = '52200') then 'Manufacturing'
            when (glb.company = 'HT' and glb.gl_account in ('45640', '45911')) or (glb.company = 'MYOP' and glb.gl_account = '52300') then 'Direct Buy Savings'
            else 'Unknown'
            end as Rebate_Category,
        -sum(glb.amount) as Rebates
    from Playground.[myop\jason.walker].gl_balance glb
        inner join Playground.[myop\jason.walker].gl_account_reporting glr
            on glb.company = glr.company
            and glb.gl_account = glr.gl_account
    where
        glr.company in ('HT', 'MYOP')
        and glr.level_1 = 'Adjustment to Margin'
        and glr.level_2 = 'Rebates'
        and glb.period between @datestart and @dateend
    group by
        glb.period,
        case when (glb.company = 'HT' and glb.gl_account = '45650') or (glb.company = 'MYOP' and glb.gl_account = '52000') then 'Pricing'
            when glb.company = 'MYOP' and glb.gl_account = '52100' then 'Wholesalers'
            when (glb.company = 'HT' and glb.gl_account = '45660') or (glb.company = 'MYOP' and glb.gl_account = '52200') then 'Manufacturing'
            when (glb.company = 'HT' and glb.gl_account in ('45640', '45911')) or (glb.company = 'MYOP' and glb.gl_account = '52300') then 'Direct Buy Savings'
            else 'Unknown'
            end    
    """.format(d1=date_end)

    rebates = gl_data = sql_retrieve.getdata1(a_driver=dict_db.get('driver'), 
                                    a_server=dict_db.get('server_2'), 
                                    a_db=dict_db.get('db_playground'), 
                                    a_sql=sql)
    
    return rebates


def get_sales_data(date_end, dict_db):
    sql = """
    declare @dateend date = '{d1}'
    declare @datestart date = dateadd("m", -11, @dateend)
    set @datestart = dateadd("d", -datepart("d", @datestart) + 1, @datestart)
    
    select
        glb.period,
        -sum(glb.amount) as Sales
    from Playground.[myop\jason.walker].gl_balance glb
        inner join Playground.[myop\jason.walker].gl_account_reporting glr
            on glb.company = glr.company
            and glb.gl_account = glr.gl_account
    where
        glr.company in ('HT', 'MYOP')
        and glr.level_1 = 'Sales'    
        and glb.period between @datestart and @dateend
    group by
        glb.period   
    """.format(d1=date_end)

    sales = sql_retrieve.getdata1(a_driver=dict_db.get('driver'), 
                                    a_server=dict_db.get('server_2'), 
                                    a_db=dict_db.get('db_playground'), 
                                    a_sql=sql)
    
    return sales

def build_report(dict_db, date_end, wb):
    rebates = get_rebate_data(date_end, dict_db)
    sales = get_sales_data(date_end, dict_db)
    date_reverse = sorted(set([x[0] for x in sales]), reverse=1)
    
    wb_cur = wb
    ws_cur = wb_cur.create_sheet('Schedule C')
    ws_cur.cell(row=1, column=1, value='Schedule C - Rebate Components')
    ws_cur.cell(row=2, column=1, value='Consolidated (HT/MYOP)')
    ws_cur.cell(row=3, 
                column=1, 
                value='For the 12 months ending {d1}'.format(
                    d1=datetime.strftime(date_end, '%B %d, %Y')))
    
    for r in range(1, 4):
        ws_cur.cell(row=r, column=1).font = Font(bold='true')
    
    date_reverse = sorted(set([x[0] for x in sales]), reverse=1)
    
    r_next = 5
    ws_cur.cell(row=r_next, column=2, value='Pricing')
    ws_cur.cell(row=r_next, column=3, value='Wholesalers')
    ws_cur.cell(row=r_next, column=4, value='Manufacturing')
    ws_cur.cell(row=r_next, column=5, value='Direct Buy')
    ws_cur.cell(row=r_next, column=6, value='Total')
    
    for c in range(2, 7):
        c1 = ws_cur.cell(row=r_next, column=c)
        c1.font = Font(bold='true', underline='single')
        c1.alignment = Alignment(horizontal='center')
        
    ws_cur.merge_cells(start_row=1, end_row=1, start_column=1, end_column=6)
    ws_cur.merge_cells(start_row=2, end_row=2, start_column=1, end_column=6)
    ws_cur.merge_cells(start_row=3, end_row=3, start_column=1, end_column=6)
    
    for r in range(1, 4):
        c1 = ws_cur.cell(row=r, column=1)
        c1.alignment = Alignment(horizontal='center')
    
    format_number = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
    border_tl = Border(top=Side(style='thin'),
                       left=Side(style='thin'))
    border_t = Border(top=Side(style='thin'))
    border_tr = Border(top=Side(style='thin'),
                       right=Side(style='thin'))
    border_bl = Border(bottom=Side(style='thin'),
                       left=Side(style='thin'))
    border_b = Border(bottom=Side(style='thin'))
    border_br = Border(bottom=Side(style='thin'),
                       right=Side(style='thin'))                                        
    
    r_next += 2    
        
    for d in date_reverse:
        month_sales = sum([s[1] for s in sales if s[0] == d])
                
        ws_cur.cell(row=r_next, column=1, 
                    value = datetime.strftime(d, '%B %Y'))        
        
        pricing = sum([r[2] for r in rebates 
                             if r[0] == d 
                             if r[1] == 'Pricing'])
        c1 = ws_cur.cell(row=r_next, column=2, 
                    value = pricing)
        c1.border = border_tl
        
        wholesaler = sum([r[2] for r in rebates 
                             if r[0] == d 
                             if r[1] == 'Wholesalers'])
        ws_cur.cell(row=r_next, column=3, 
                    value = wholesaler)        
        
        manufacturing = sum([r[2] for r in rebates 
                             if r[0] == d 
                             if r[1] == 'Manufacturing'])
        ws_cur.cell(row=r_next, column=4, 
                    value = manufacturing)
        
        direct_buy = sum([r[2] for r in rebates 
                             if r[0] == d 
                             if r[1] == 'Direct Buy Savings'])                
        ws_cur.cell(row=r_next, column=5, 
                    value = direct_buy)
        
        formula_total = '=sum(B{r1}:E{r1})'.format(r1=r_next)
        c1 = ws_cur.cell(row=r_next, column=6, value=formula_total)
        c1.border = border_tr
    
        for c in range(2, 7):
            ws_cur.cell(row=r_next, column=c).number_format = format_number
        
        for c in range(3, 6):
            c1 = ws_cur.cell(row=r_next, column=c).border = border_t            
        
        
        
        pricing_ratio = pricing / month_sales if month_sales else 0
        wholesaler_ratio = wholesaler / month_sales if month_sales else 0
        manufacturing_ratio = manufacturing / month_sales if month_sales else 0
        direct_buy_ratio = direct_buy / month_sales if month_sales else 0
        
        ws_cur.cell(row=r_next + 1, column=1, value='% of Revenue')
        ws_cur.cell(row=r_next + 1, column=2, value=pricing_ratio)
        ws_cur.cell(row=r_next + 1, column=3, value=wholesaler_ratio)
        ws_cur.cell(row=r_next + 1, column=4, value=manufacturing_ratio)
        ws_cur.cell(row=r_next + 1, column=5, value=direct_buy_ratio)
        formula_total = '=sum(B{r1}:E{r1})'.format(r1=r_next + 1)
        ws_cur.cell(row=r_next + 1, column=6, value=formula_total)
        
        for c in range(2, 7):
            ws_cur.cell(row=r_next + 1, column=c).number_format = '0.0%'
                
        ws_cur.cell(row=r_next + 1, column=2).border = border_bl
                
        for c in range(3, 6):
            ws_cur.cell(row=r_next + 1, column=c).border = border_b
        
        ws_cur.cell(row=r_next + 1, column=6).border = border_br
        
        r_next += 3
    
    #Total
    ws_cur.cell(row=r_next, column=1, value='Rolling 12 Month Total')
    
    for c in range(2, 7):
        r = 3
        cat_total = '='
        
        for x in range(1, 13):
            cat_total = cat_total + '{c1}{r1}+'.format(
                c1=dcc.get(c), r1=r_next - r)
            r += 3 
        
        cat_total = cat_total[:-1]
        ws_cur.cell(row=r_next, column=c, value=cat_total)
    
    ws_cur.cell(row=r_next + 1, column=1, value='% of Revenue')
    pricing = sum([r[2] for r in rebates
                         if r[1] == 'Pricing'])    
    wholesaler = sum([r[2] for r in rebates
                         if r[1] == 'Wholesalers'])
    manufacturing = sum([r[2] for r in rebates
                         if r[1] == 'Manufacturing'])    
    direct_buy = sum([r[2] for r in rebates
                         if r[1] == 'Direct Buy Savings'])
    rebate_total = sum([r[2] for r in rebates])
    sales_total = sum([s[1] for s in sales])
    pricing_ratio = pricing / sales_total if sales_total else 0
    wholesaler_ratio = wholesaler / sales_total if sales_total else 0
    manufacturing_ratio = manufacturing / sales_total if sales_total else 0
    direct_buy_ratio = direct_buy / sales_total if sales_total else 0    
    rebate_total_ratio = rebate_total / sales_total if sales_total else 0    
    
    ws_cur.cell(row=r_next + 1, column=2, value=pricing_ratio)
    ws_cur.cell(row=r_next + 1, column=3, value=wholesaler_ratio)
    ws_cur.cell(row=r_next + 1, column=4, value=manufacturing_ratio)
    ws_cur.cell(row=r_next + 1, column=5, value=direct_buy_ratio)
    ws_cur.cell(row=r_next + 1, column=6, value=rebate_total_ratio)
    
    for c in range(2, 7):
        ws_cur.cell(row=r_next, column=c).number_format = format_number    
        ws_cur.cell(row=r_next + 1, column=c).number_format = '0.0%'
    
    ws_cur.cell(row=r_next, column=2).border = border_tl
    ws_cur.cell(row=r_next + 1, column=2).border = border_bl
    
    for c in range(3, 6):
        ws_cur.cell(row=r_next, column=c).border = border_t
        ws_cur.cell(row=r_next + 1, column=c).border = border_b
    
    ws_cur.cell(row=r_next, column=6).border = border_tr
    ws_cur.cell(row=r_next + 1, column=6).border = border_br
    
    
    #Page Setup
    ws_cur.column_dimensions[dcc.get(1)].width = 22
    
    for c in range(2, 7):
        ws_cur.column_dimensions[dcc.get(c)].width = 15    
    
    ws_cur.page_setup.orientation = ws_cur.ORIENTATION_PORTRAIT
    ws_cur.page_setup.paper_size = ws_cur.PAPERSIZE_TABLOID
    ws_cur.page_setup.fitToPage = True
    ws_cur.page_setup.fitToHeight = 1
    ws_cur.page_setup.fitToWidth = 1
    ws_cur.print_options.horizontalCentered = True
    ws_cur.add_print_title(5)
    ws_cur.page_margins = PageMargins(left=.5, right=.5, top=.5, bottom=.5,
                                      footer=.5)
    
    #Freeze Panes
    c1 = ws_cur.cell(row=6, column=1)
    ws_cur.freeze_panes = c1    
    
    
    
    
    
    
    
    
    
    
    return wb_cur




















if __name__ == "__main__":
    get_rebate_data()
    get_sales_data()
    build_report()
    