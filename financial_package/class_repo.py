import pypyodbc
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles.borders import Border
from openpyxl.styles.borders import Side
from excel_column_number import dict_col_converter as dcc
from datetime import datetime

class SQLExchange:
    
    def __init__(self, driver, server, db):
        self.driver = driver
        self.server = server
        self.db = db            
    
        self.connection_string = '''Driver={d1};
                                Server={s1};
                                Database={db1};
                                Trusted_Connection=Yes;
                            '''.format(d1=self.driver, 
                                       s1=self.server, 
                                       db1=self.db)
        
        
    def sql_retrieve(self, sql):        
        connection = pypyodbc.connect(self.connection_string) 
        cur = connection.cursor()
        sql_result = cur.execute(sql).fetchall()
        cur.close()
        connection.close
        return sql_result    
                        
    def sql_no_retrieve(self, sql):
        connection = pypyodbc.connect(self.connection_string)
        cur = connection.cursor()
        cur.execute(sql)
        cur.commit()
        cur.close()
        connection.close
        return None
    
    
class ExcelWidget:    
    
    def __init__(self, wb):
        self.wb = wb
        
    def arial_8_bold_center(self, cell):
        wb_new = self.wb             
        c1 = cell
        c1.font = Font(name='Arial', size=8, bold='true')
        c1.alignment = Alignment(horizontal='center')        
        return wb_new
    
    def header_blue_arial_8(self, cell):
        wb_new = self.wb        
        c1 = cell
        c1.fill = PatternFill(start_color='1e90ff',
                              end_color='1e90ff',
                              fill_type='solid')
        c1.font = Font(bold='true', color='f8f8ff', 
                       name='Arial', size=8)
        c1.alignment = Alignment(horizontal='center')
        return wb_new

    def arial_8_yellow_bold_number0(self, cell):
        wb_new = self.wb        
        c1 = cell
        c1.fill = PatternFill(start_color='ffd700',
                              end_color='ffd700',
                              fill_type='solid')
        c1.font = Font(bold='true', name='Arial', size=8)
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'
        c1.alignment = Alignment(horizontal='center')
        return wb_new

    
    def arial_8_bold(self, cell):
        wb_new = self.wb             
        c1 = cell
        c1.font = Font(name='Arial', size=8, bold='true')                
        return wb_new        

    def arial_8_italic(self, cell):
        wb_new = self.wb             
        c1 = cell
        c1.font = Font(name='Arial', size=8, italic='true')                
        return wb_new


    def arial_8(self, cell):
        wb_new = self.wb             
        c1 = cell
        c1.font = Font(name='Arial', size=8)                
        return wb_new
            
    def arial_8_bold_number0(self, cell):
        wb_new = self.wb             
        c1 = cell
        c1.font = Font(name='Arial', size=8, bold='true')
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'                
        return wb_new
        
    def row_total(self, colstart, colend, rcur):
        formula1 = "=SUM({c1}{r1}:{c2}{r1})".format(c1=dcc.get(colstart), 
                                                c2=dcc.get(colend), 
                                                r1=rcur)
        return formula1
    
    def col_total(self, rowstart, rowend, colcur):
        formula1 = "=SUM({c1}{r1}:{c1}{r2})".format(c1=dcc.get(colcur),
                                                  r1=rowstart,
                                                  r2=rowend)
        return formula1
    
    def sum_2(self, c1, c2):        
        formula1 = "={c1}{r1}+{c2}{r2}".format(c1=c1.column,
                                               c2=c2.column,
                                                  r1=c1.row,
                                                  r2=c2.row)
        return formula1    

    def sum_3(self, c1, c2, c3):        
        formula1 = "={c1}{r1}+{c2}{r2}+{c3}{r3}".format(c1=c1.column,
                                               r1=c1.row,
                                               c2=c2.column,                                                  
                                               r2=c2.row,
                                               c3=c3.column,
                                               r3=c3.row)
        return formula1    

    def subtract_2(self, c1, c2):        
        formula1 = "={c1}{r1}-{c2}{r2}".format(c1=c1.column,
                                               c2=c2.column,
                                                  r1=c1.row,
                                                  r2=c2.row)
        return formula1
    
    def ebitda_calc(self, agm, opex, oth_income):        
        formula1 = "={c1}{r1}-{c2}{r2}+{c3}{r3}".format(c1=agm.column,
                                               r1=agm.row,
                                               c2=opex.column,                                                  
                                               r2=opex.row,
                                               c3=oth_income.column,
                                               r3=oth_income.row)
        return formula1    

    def net_income_before_tax(self, ebitda, ida_start, ida_end):        
        formula1 = "={c1}{r1}-sum({c2}{r2}:{c3}{r3})".format(c1=ebitda.column,
                                               r1=ebitda.row,
                                               c2=ida_start.column,                                                  
                                               r2=ida_start.row,
                                               c3=ida_end.column,
                                               r3=ida_end.row)
        return formula1
    

    
    def divide_2(self, divisor, dividend):
        formula1 = "=IFERROR({c1}{r1}/{c2}{r2},0)".format(c1=divisor.column,
                                                          c2=dividend.column,
                                                          r1=divisor.row,
                                                          r2=dividend.row)
        return formula1
    
    def excluded_expense_inverse(self, excluded_expense):        
        formula1 = "=-{c1}{r1}".format(c1=excluded_expense.column,
                                               r1=excluded_expense.row)
        return formula1                                                          
    
    def border_tb(self, cell):
        wb_new = self.wb
        c1 = cell
        border_tb = Border(top=Side(style='thin'), 
                           bottom=Side(style='thin'))
        c1.border = border_tb
        return wb_new

    def border_tb2(self, cell):
        wb_new = self.wb
        c1 = cell
        border_tb = Border(top=Side(style='thin'), 
                           bottom=Side(style='double'))
        c1.border = border_tb
        return wb_new
    
    def page_setup_portrait_1X1(self, ws, print_rows):
        wb_new = self.wb
        ws_cur = ws
        ws_cur.page_setup.orientation = ws_cur.ORIENTATION_PORTRAIT
        ws_cur.page_setup.paper_size = ws_cur.PAPERSIZE_TABLOID
        ws_cur.page_setup.fitToPage = True
        ws_cur.page_setup.fitToHeight = 1
        ws_cur.page_setup.fitToWidth = 1
        ws_cur.print_options.horizontalCentered = True
        ws_cur.add_print_title(print_rows)
        ws_cur.page_margins = PageMargins(left=.5, right=.5, top=.5, bottom=.5,
                                          footer=.5)        
        date_cur = datetime.strftime(datetime.today(), 
                              "%a {dt.month}/{dt.day}/%Y %I:%M %p".format(
                                  dt=datetime.today()))
        
        ws_cur.oddFooter.right.text = date_cur
        ws_cur.oddFooter.right.size = 8
        ws_cur.oddFooter.right.font = 'Arial'
        return wb_new

    def page_setup_landscape_1X1(self, ws, print_rows):
        wb_new = self.wb
        ws_cur = ws
        ws_cur.page_setup.orientation = ws_cur.ORIENTATION_LANDSCAPE
        ws_cur.page_setup.paper_size = ws_cur.PAPERSIZE_TABLOID
        ws_cur.page_setup.fitToPage = True
        ws_cur.page_setup.fitToHeight = 1
        ws_cur.page_setup.fitToWidth = 1
        ws_cur.print_options.horizontalCentered = True
        ws_cur.add_print_title(print_rows)
        ws_cur.page_margins = PageMargins(left=.5, right=.5, top=.5, bottom=.5,
                                          footer=.5)        
        date_cur = datetime.strftime(datetime.today(), 
                              "%a {dt.month}/{dt.day}/%Y %I:%M %p".format(
                                  dt=datetime.today()))
        
        ws_cur.oddFooter.right.text = date_cur
        ws_cur.oddFooter.right.size = 8
        ws_cur.oddFooter.right.font = 'Arial'
        return wb_new
        
    def arial_8_number0(self, cell):
        wb_new = self.wb             
        c1 = cell
        c1.font = Font(name='Arial', size=8)
        c1.number_format = '_(* #,##0_);_(* (#,##0);_(* ""-""??_);_(@_)'                
        return wb_new        

    def arial_8_italic_percent1(self, cell):
        wb_new = self.wb             
        c1 = cell
        c1.font = Font(name='Arial', size=8, italic='true')
        c1.number_format ='0.0%'                
        return wb_new        